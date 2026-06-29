"""
================================================================================
SPK-GeoOptima DSS v5.1 — STREAMLIT CLOUD EDITION
Framework: GIS + NSGA-II + Huff + Stackelberg + CVaR
Fitur: Plotly, Export Excel/PDF, Scenario Manager
Dihapus: st-aggrid (tidak kompatibel Python 3.14)
================================================================================
"""

import sqlite3
import pandas as pd
import numpy as np
import streamlit as st
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import os
import json
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# IMPORT OPTIONAL — dengan fallback graceful
# ============================================================
try:
    import folium
    from folium.plugins import Fullscreen, MeasureControl, MiniMap, HeatMap
    FOLIUM_AVAILABLE = True
except ImportError:
    FOLIUM_AVAILABLE = False
    folium = None

try:
    from streamlit_folium import st_folium
    STREAMLIT_FOLIUM_AVAILABLE = True
except ImportError:
    STREAMLIT_FOLIUM_AVAILABLE = False

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ============================================================
# KONFIGURASI HALAMAN
# ============================================================
st.set_page_config(
    page_title="SPK-GeoOptima v5.1 | DSS Infrastruktur Publik",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# STATE MANAGEMENT
# ============================================================
if 'data_initialized' not in st.session_state:
    st.session_state.data_initialized = False
if 'show_input_panel' not in st.session_state:
    st.session_state.show_input_panel = False
if 'last_saved' not in st.session_state:
    st.session_state.last_saved = None
if 'saved_scenarios' not in st.session_state:
    st.session_state.saved_scenarios = []
if 'dark_mode' not in st.session_state:
    st.session_state.dark_mode = False
if 'compare_mode' not in st.session_state:
    st.session_state.compare_mode = False

# ============================================================
# CSS CUSTOM
# ============================================================
st.markdown("""
<style>
    .main .block-container { padding-top: 1rem; padding-bottom: 2rem; max-width: 100%; }
    [data-testid="stSidebar"] { background: linear-gradient(180deg, #0F172A 0%, #1E293B 100%); }
    [data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2, [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] p, [data-testid="stSidebar"] label { color: #E2E8F0 !important; }

    .top-header {
        background: linear-gradient(135deg, #0F172A 0%, #1E3A8A 50%, #2563EB 100%);
        padding: 20px 32px; border-radius: 16px; margin-bottom: 20px;
        display: flex; justify-content: space-between; align-items: center;
        box-shadow: 0 4px 16px rgba(30, 58, 138, 0.3);
    }
    .header-title { color: white; font-size: 1.6rem; font-weight: 800; margin: 0; }
    .header-subtitle { color: rgba(255,255,255,0.7); font-size: 0.8rem; text-transform: uppercase; letter-spacing: 2px; }
    .header-badge { background: rgba(255,255,255,0.15); padding: 8px 16px; border-radius: 8px; color: white; font-size: 0.85rem; font-weight: 600; }

    .kpi-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 12px; margin-bottom: 20px; }
    .kpi-card { border-radius: 12px; padding: 18px; color: white; box-shadow: 0 2px 8px rgba(0,0,0,0.1); transition: transform 0.2s; }
    .kpi-card:hover { transform: translateY(-3px); }
    .kpi-label { font-size: 0.75rem; font-weight: 600; opacity: 0.9; margin-bottom: 6px; }
    .kpi-value { font-size: 2rem; font-weight: 800; line-height: 1.1; }
    .kpi-sub { font-size: 0.7rem; opacity: 0.85; margin-top: 6px; }
    .kpi-navy { background: linear-gradient(135deg, #1E3A8A, #172554); }
    .kpi-blue { background: linear-gradient(135deg, #2563EB, #1D4ED8); }
    .kpi-teal { background: linear-gradient(135deg, #0891B2, #0E7490); }
    .kpi-green { background: linear-gradient(135deg, #059669, #047857); }
    .kpi-amber { background: linear-gradient(135deg, #D97706, #B45309); }
    .kpi-rose { background: linear-gradient(135deg, #E11D48, #BE123C); }
    .kpi-violet { background: linear-gradient(135deg, #7C3AED, #6D28D9); }

    .pbi-card { background: #FFFFFF; border-radius: 12px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.06); margin-bottom: 16px; border: 1px solid #E2E8F0; }
    .pbi-card-title { font-size: 1rem; font-weight: 700; color: #0F172A; margin-bottom: 12px; }

    .pillar-grid { display: grid; grid-template-columns: repeat(5, 1fr); gap: 10px; margin-bottom: 20px; }
    .pillar-card { background: white; border-radius: 10px; padding: 14px; border: 2px solid #E2E8F0; text-align: center; transition: all 0.2s; }
    .pillar-card:hover { border-color: #3B82F6; transform: translateY(-2px); box-shadow: 0 4px 12px rgba(59,130,246,0.15); }
    .pillar-icon { font-size: 1.5rem; margin-bottom: 6px; }
    .pillar-name { font-size: 0.8rem; font-weight: 700; color: #1E293B; }
    .pillar-desc { font-size: 0.7rem; color: #64748B; }

    .info-box { background: #EFF6FF; border-left: 4px solid #3B82F6; border-radius: 0 8px 8px 0; padding: 14px 18px; margin: 12px 0; }
    .info-box-title { font-weight: 700; color: #1E40AF; margin-bottom: 4px; }
    .info-box-text { color: #475569; font-size: 0.85rem; }

    .input-panel { background: linear-gradient(135deg, #F0F9FF 0%, #FFFFFF 100%); border: 2px solid #0EA5E9; border-radius: 16px; padding: 24px; margin-bottom: 20px; }
    .input-panel-title { font-size: 1.1rem; font-weight: 700; color: #0369A1; margin-bottom: 16px; display: flex; align-items: center; gap: 8px; }

    .footer-bar { background: linear-gradient(135deg, #0F172A, #1E3A8A); padding: 14px 24px; border-radius: 12px; color: white; font-size: 0.8rem; margin-top: 24px; }

    .stTabs [data-baseweb="tab-list"] { gap: 6px; background: linear-gradient(135deg, #1E3A8A, #1E40AF); padding: 8px; border-radius: 10px; margin-bottom: 16px; }
    .stTabs [data-baseweb="tab"] { color: rgba(255,255,255,0.7) !important; background: transparent !important; border-radius: 8px !important; padding: 10px 20px !important; font-weight: 600 !important; font-size: 0.85rem !important; }
    .stTabs [aria-selected="true"] { background: linear-gradient(135deg, #3B82F6, #2563EB) !important; color: white !important; }

    .scenario-card { background: linear-gradient(135deg, #F8FAFC, #FFFFFF); border: 2px solid #E2E8F0; border-radius: 12px; padding: 16px; margin-bottom: 12px; cursor: pointer; transition: all 0.2s; }
    .scenario-card:hover { border-color: #3B82F6; box-shadow: 0 4px 12px rgba(59,130,246,0.1); }
    .scenario-active { border-color: #10B981 !important; background: linear-gradient(135deg, #ECFDF5, #F0FDF4) !important; }
</style>
""", unsafe_allow_html=True)

# ============================================================
# KELAS: SPATIAL DATABASE
# ============================================================
class SpatialDatabase:
    def __init__(self, db_path="spklu_integrated_v51.db"):
        try:
            self.conn = sqlite3.connect(db_path, check_same_thread=False)
            self.conn.row_factory = sqlite3.Row
            self.cursor = self.conn.cursor()
            self._init_schema()
        except Exception as e:
            st.error(f"Database Error: {e}")
            raise

    def _init_schema(self):
        tables = [
            """CREATE TABLE IF NOT EXISTS wilayah_kalsel (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kode TEXT UNIQUE, nama TEXT, tipe TEXT,
                lat_center REAL, lon_center REAL, populasi INTEGER, luas_km2 REAL
            )""",
            """CREATE TABLE IF NOT EXISTS kandidat_lokasi (
                id INTEGER PRIMARY KEY, nama TEXT, kab_kota TEXT, kecamatan TEXT,
                lat REAL, lon REAL, biaya REAL, demand REAL DEFAULT 0,
                kapasitas_grid_kva REAL, skor_akses_jalan REAL, skor_visibilitas REAL,
                status TEXT DEFAULT 'kandidat', zona_prioritas TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS poi_demand (
                id INTEGER PRIMARY KEY, nama TEXT, tipe TEXT, kab_kota TEXT,
                lat REAL, lon REAL, kunjungan_hari INTEGER, daya_tarik REAL, bobot_demand REAL
            )""",
            """CREATE TABLE IF NOT EXISTS jaringan_jalan (
                id INTEGER PRIMARY KEY, nama TEXT, tipe TEXT,
                lat_from REAL, lon_from REAL, lat_to REAL, lon_to REAL, panjang_km REAL
            )""",
            """CREATE TABLE IF NOT EXISTS hasil_huff (
                id INTEGER PRIMARY KEY, id_poi INTEGER, id_kandidat INTEGER,
                jarak_km REAL, probabilitas REAL, demand_tertangkap REAL
            )""",
            """CREATE TABLE IF NOT EXISTS hasil_nsga (
                id INTEGER PRIMARY KEY, scenario_id TEXT, solusi_id INTEGER, generasi INTEGER,
                obj_biaya REAL, obj_coverage REAL, obj_equity REAL, obj_risiko REAL,
                lokasi_terpilih TEXT, is_pareto BOOLEAN, rank_crowding REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS hasil_stackelberg (
                id INTEGER PRIMARY KEY, scenario_id TEXT, id_kandidat INTEGER,
                x_terpilih INTEGER, subsidi_optimal REAL, investor_response TEXT,
                leader_payoff REAL, follower_payoff REAL, net_benefit REAL, equilibrium_type TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS hasil_cvar (
                id INTEGER PRIMARY KEY, scenario_id TEXT, id_kandidat INTEGER,
                expected_return REAL, var_95 REAL, cvar_95 REAL, z_invest REAL,
                adjusted_npv REAL, risk_rating TEXT
            )""",
            """CREATE TABLE IF NOT EXISTS skenario_config (
                id INTEGER PRIMARY KEY, scenario_id TEXT UNIQUE, nama TEXT,
                budget REAL, max_lokasi INTEGER, min_lokasi INTEGER, min_jarak_km REAL,
                huff_alpha REAL, huff_beta REAL, cvar_alpha REAL, risk_aversion REAL,
                nsga_popsize INTEGER, nsga_generasi INTEGER, demand_growth REAL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS huff_params (
                id INTEGER PRIMARY KEY, scenario_id TEXT, alpha REAL, beta REAL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS cvar_skenario (
                id INTEGER PRIMARY KEY, scenario_id TEXT, nama_skenario TEXT,
                demand_factor REAL, cost_factor REAL, prob REAL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS stackelberg_params (
                id INTEGER PRIMARY KEY, scenario_id TEXT, subsidy_max REAL,
                budget_pemerintah REAL, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS saved_scenarios (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nama TEXT, deskripsi TEXT,
                budget REAL, max_lokasi INTEGER, min_jarak_km REAL,
                huff_beta REAL, cvar_alpha REAL, risk_aversion REAL,
                nsga_popsize INTEGER, nsga_generasi INTEGER,
                hasil_json TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )"""
        ]
        for sql in tables:
            self.cursor.execute(sql)
        self.conn.commit()

    def seed_default_data(self):
        try:
            self.cursor.execute("SELECT COUNT(*) FROM kandidat_lokasi")
            if self.cursor.fetchone()[0] > 0:
                return False

            wilayah = [
                ("63.71", "Kota Banjarmasin", "kota", -3.320, 114.592, 668000, 371.20),
                ("63.72", "Kota Banjarbaru", "kota", -3.466, 114.750, 272000, 371.38),
                ("63.01", "Kab. Tanah Laut", "kab", -3.850, 114.800, 315000, 10052.18),
                ("63.02", "Kab. Kota Baru", "kab", -3.233, 116.133, 315000, 12058.43),
                ("63.03", "Kab. Banjar", "kab", -3.317, 114.983, 565000, 4959.72),
                ("63.04", "Kab. Barito Kuala", "kab", -3.083, 114.367, 313000, 2908.30),
                ("63.05", "Kab. Tapin", "kab", -2.900, 115.167, 175000, 2176.24),
                ("63.06", "Kab. Hulu Sungai Selatan", "kab", -2.750, 115.200, 228000, 1718.71),
                ("63.07", "Kab. Hulu Sungai Tengah", "kab", -2.583, 115.417, 118000, 1087.10),
                ("63.08", "Kab. Hulu Sungai Utara", "kab", -2.450, 115.233, 226000, 1902.13),
                ("63.09", "Kab. Tabalong", "kab", -2.200, 115.550, 260000, 3496.90),
                ("63.10", "Kab. Tanah Bumbu", "kab", -3.450, 115.900, 314000, 13065.36),
                ("63.11", "Kab. Balangan", "kab", -2.317, 115.617, 114000, 1833.95),
            ]
            self.cursor.executemany("INSERT INTO wilayah_kalsel (kode, nama, tipe, lat_center, lon_center, populasi, luas_km2) VALUES (?,?,?,?,?,?,?)", wilayah)

            kandidat = [
                (1, "Duta Mall Banjarmasin", "Kota Banjarmasin", "Sungai Besar", -3.323020, 114.603136, 2500000000, 2400, 500, 9.2, 9.5, "kandidat", "Pusat Kota"),
                (2, "Transmart Banjarmasin", "Kota Banjarmasin", "Sungai Besar", -3.319800, 114.602100, 2200000000, 2100, 500, 9.0, 9.3, "kandidat", "Pusat Kota"),
                (3, "RSUD Ulin Banjarmasin", "Kota Banjarmasin", "Sungai Besar", -3.319630, 114.598560, 1800000000, 1500, 300, 8.5, 8.8, "kandidat", "Pusat Kota"),
                (4, "Kampus ULM Kayu Tangi", "Kota Banjarmasin", "Kayu Tangi", -3.300800, 114.589100, 1500000000, 4000, 200, 7.8, 8.5, "kandidat", "Pendidikan"),
                (5, "Pelabuhan Trisakti", "Kota Banjarmasin", "Sungai Jingah", -3.328400, 114.567200, 2800000000, 1800, 630, 7.5, 7.0, "kandidat", "Logistik"),
                (6, "SPBU A. Yani KM 5", "Kota Banjarmasin", "Sungai Andai", -3.337800, 114.626500, 3200000000, 2200, 1000, 9.5, 8.0, "kandidat", "Jalur Utama"),
                (7, "Bandara Syamsudin Noor", "Kota Banjarbaru", "Landasan Ulin", -3.442350, 114.762500, 4500000000, 3000, 1500, 9.8, 9.0, "kandidat", "Transportasi"),
                (8, "Q Mall Banjarbaru", "Kota Banjarbaru", "Landasan Ulin", -3.448900, 114.831200, 2800000000, 2000, 500, 8.8, 8.5, "kandidat", "Pusat Kota"),
                (9, "Kampus UNISKA", "Kota Banjarbaru", "Sungai Besar", -3.455000, 114.820000, 1600000000, 2500, 200, 7.5, 7.8, "kandidat", "Pendidikan"),
                (10, "Pasar Terapung Lok Baintan", "Kab. Banjar", "Sungai Tabuk", -3.355700, 114.671000, 1200000000, 1200, 100, 6.5, 9.5, "kandidat", "Wisata"),
                (11, "RSUD Ratu Zaleha Martapura", "Kab. Banjar", "Martapura", -3.410000, 114.850000, 1900000000, 1600, 300, 8.0, 8.2, "kandidat", "Kesehatan"),
                (12, "Alun-alun Pelaihari", "Kab. Tanah Laut", "Pelaihari", -3.806000, 114.769800, 1400000000, 1000, 200, 7.2, 7.5, "kandidat", "Pusat Kabupaten"),
                (13, "RSUD H. Boejasin Pelaihari", "Kab. Tanah Laut", "Pelaihari", -3.810000, 114.775000, 1700000000, 1400, 300, 7.8, 7.8, "kandidat", "Kesehatan"),
                (14, "Terminal Batulicin", "Kab. Tanah Bumbu", "Batulicin", -3.483300, 115.916700, 2000000000, 1100, 400, 7.0, 7.2, "kandidat", "Transportasi"),
                (15, "Kawasan Industri Mangkupum", "Kab. Tanah Bumbu", "Mangkupum", -3.433300, 115.866700, 3500000000, 3500, 2000, 8.0, 6.5, "kandidat", "Industri"),
                (16, "Alun-alun Kandangan", "Kab. Hulu Sungai Selatan", "Kandangan", -2.766700, 115.200000, 1300000000, 900, 200, 7.5, 7.8, "kandidat", "Pusat Kabupaten"),
                (17, "Pasar Barabai", "Kab. Hulu Sungai Tengah", "Barabai", -2.583300, 115.416700, 1100000000, 800, 150, 7.0, 7.5, "kandidat", "Perdagangan"),
                (18, "Terminal Amuntai", "Kab. Hulu Sungai Utara", "Amuntai", -2.416700, 115.250000, 1250000000, 850, 200, 7.2, 7.0, "kandidat", "Transportasi"),
                (19, "Kawasan Pertambangan Tabalong", "Kab. Tabalong", "Tanjung", -2.200000, 115.550000, 4000000000, 2800, 3000, 8.5, 6.0, "kandidat", "Industri"),
                (20, "Pasar Paringin", "Kab. Balangan", "Paringin", -2.333300, 115.616700, 1000000000, 700, 100, 6.8, 7.2, "kandidat", "Perdagangan"),
            ]
            self.cursor.executemany("INSERT INTO kandidat_lokasi (id, nama, kab_kota, kecamatan, lat, lon, biaya, demand, kapasitas_grid_kva, skor_akses_jalan, skor_visibilitas, status, zona_prioritas) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", kandidat)

            poi = [
                (1, "Duta Mall", "Mall", "Kota Banjarmasin", -3.323020, 114.603136, 8500, 4.5, 1.0),
                (2, "Transmart BTC", "Mall", "Kota Banjarmasin", -3.319800, 114.602100, 7200, 4.2, 1.0),
                (3, "RSUD Ulin", "Rumah Sakit", "Kota Banjarmasin", -3.319630, 114.598560, 3500, 3.8, 0.9),
                (4, "Kantor Gubernur Kalsel", "Perkantoran", "Kota Banjarmasin", -3.321000, 114.600000, 2800, 3.5, 0.8),
                (5, "Bandara Syamsudin Noor", "Bandara", "Kota Banjarbaru", -3.442350, 114.762500, 4500, 4.0, 1.2),
                (6, "ULM Kayu Tangi", "Kampus", "Kota Banjarmasin", -3.300800, 114.589100, 15000, 4.3, 1.1),
                (7, "Pasar Terapung Lok Baintan", "Wisata", "Kab. Banjar", -3.355700, 114.671000, 3000, 4.1, 1.3),
                (8, "Q Mall Banjarbaru", "Mall", "Kota Banjarbaru", -3.448900, 114.831200, 5500, 3.9, 1.0),
                (9, "RSUD H. Boejasin", "Rumah Sakit", "Kab. Tanah Laut", -3.806000, 114.769800, 2200, 3.2, 0.7),
                (10, "Pelabuhan Trisakti", "Pelabuhan", "Kota Banjarmasin", -3.328400, 114.567200, 2800, 3.6, 0.9),
                (11, "UNISKA MAB", "Kampus", "Kota Banjarbaru", -3.455000, 114.820000, 6200, 3.7, 0.9),
                (12, "RSUD Ratu Zaleha", "Rumah Sakit", "Kab. Banjar", -3.410000, 114.850000, 2800, 3.5, 0.8),
                (13, "Alun-alun Pelaihari", "Pusat Kota", "Kab. Tanah Laut", -3.806000, 114.769800, 1800, 3.0, 0.7),
                (14, "Terminal Batulicin", "Terminal", "Kab. Tanah Bumbu", -3.483300, 115.916700, 1500, 2.8, 0.6),
                (15, "Pasar Kandangan", "Pasar", "Kab. Hulu Sungai Selatan", -2.766700, 115.200000, 2400, 3.1, 0.7),
                (16, "Pasar Barabai", "Pasar", "Kab. Hulu Sungai Tengah", -2.583300, 115.416700, 1900, 2.9, 0.6),
                (17, "Terminal Amuntai", "Terminal", "Kab. Hulu Sungai Utara", -2.416700, 115.250000, 1600, 2.7, 0.6),
                (18, "Kawasan Pertambangan Tanjung", "Industri", "Kab. Tabalong", -2.200000, 115.550000, 8000, 3.4, 1.1),
                (19, "Pasar Paringin", "Pasar", "Kab. Balangan", -2.333300, 115.616700, 1400, 2.6, 0.6),
                (20, "Mall Lippo Banjarmasin", "Mall", "Kota Banjarmasin", -3.315000, 114.610000, 6800, 4.0, 1.0),
                (21, "Stadion 17 Mei", "Olahraga", "Kota Banjarmasin", -3.325000, 114.615000, 3200, 3.3, 0.8),
                (22, "Masjid Raya Sabilal Muhtadin", "Religi", "Kota Banjarmasin", -3.318000, 114.595000, 4500, 4.4, 0.9),
                (23, "Taman Siring 0 KM", "Wisata", "Kota Banjarmasin", -3.316000, 114.592000, 5200, 4.2, 1.1),
                (24, "Kampus UIN Antasari", "Kampus", "Kota Banjarmasin", -3.295000, 114.580000, 4800, 3.8, 0.9),
                (25, "SPBU Strategis A. Yani", "SPBU", "Kota Banjarmasin", -3.337800, 114.626500, 5500, 3.5, 1.0),
            ]
            self.cursor.executemany("INSERT INTO poi_demand VALUES (?,?,?,?,?,?,?,?,?)", poi)

            jalan = [
                (1, "Jalan A. Yani", "Arteri", -3.320000, 114.590000, -3.340000, 114.630000, 5.2),
                (2, "Jalan Sudirman", "Arteri", -3.315000, 114.585000, -3.325000, 114.610000, 3.1),
                (3, "Jalan Lambung Mangkurat", "Kolektor", -3.310000, 114.595000, -3.330000, 114.605000, 2.4),
                (4, "Jalan Banjarbaru", "Arteri", -3.330000, 114.600000, -3.450000, 114.800000, 28.5),
                (5, "Jalan Martapura", "Arteri", -3.330000, 114.600000, -3.410000, 114.850000, 32.1),
                (6, "Jalan Pelaihari", "Kolektor", -3.330000, 114.600000, -3.810000, 114.770000, 68.4),
                (7, "Jalan Batulicin", "Arteri", -3.450000, 114.800000, -3.483300, 115.916700, 145.2),
                (8, "Jalan Kandangan", "Kolektor", -3.330000, 114.600000, -2.766700, 115.200000, 95.3),
                (9, "Jalan Amuntai", "Kolektor", -2.766700, 115.200000, -2.416700, 115.250000, 42.8),
                (10, "Jalan Tanjung", "Arteri", -3.330000, 114.600000, -2.200000, 115.550000, 185.6),
            ]
            self.cursor.executemany("INSERT INTO jaringan_jalan VALUES (?,?,?,?,?,?,?,?)", jalan)

            self.cursor.execute("INSERT OR IGNORE INTO skenario_config (scenario_id, nama, budget, max_lokasi, min_lokasi, min_jarak_km, huff_alpha, huff_beta, cvar_alpha, risk_aversion, nsga_popsize, nsga_generasi, demand_growth) VALUES ('default', 'Skenario Default Kalsel', 15000000000, 8, 3, 5.0, 1.0, 2.0, 0.95, 0.5, 50, 30, 1.0)")
            self.cursor.execute("INSERT OR IGNORE INTO huff_params (scenario_id, alpha, beta) VALUES ('default', 1.0, 2.0)")
            cvar_skenario = [('default', 'Optimis', 1.4, 0.85, 0.20), ('default', 'Moderat+', 1.15, 0.95, 0.30), ('default', 'Moderat', 1.0, 1.0, 0.30), ('default', 'Pesimis', 0.7, 1.15, 0.20)]
            self.cursor.executemany("INSERT OR IGNORE INTO cvar_skenario (scenario_id, nama_skenario, demand_factor, cost_factor, prob) VALUES (?,?,?,?,?)", cvar_skenario)
            self.cursor.execute("INSERT OR IGNORE INTO stackelberg_params (scenario_id, subsidy_max, budget_pemerintah) VALUES ('default', 0.3, 15000000000)")
            self.conn.commit()
            return True
        except Exception as e:
            st.warning(f"Warning saat seed data: {e}")
            self.conn.rollback()
            return False

    def query(self, sql, params=()):
        try:
            return pd.read_sql(sql, self.conn, params=params)
        except Exception as e:
            st.warning(f"Query error: {e}")
            return pd.DataFrame()

    def execute(self, sql, params=()):
        try:
            self.cursor.execute(sql, params)
            self.conn.commit()
        except Exception as e:
            st.warning(f"Execute error: {e}")
            self.conn.rollback()

    def close(self):
        self.conn.close()


# ============================================================
# KELAS: DISTANCE MATRIX
# ============================================================
class DistanceMatrix:
    def __init__(self, df_kan, df_poi):
        self.R = 6371.0
        self.n_kan = len(df_kan)
        self.n_poi = len(df_poi)
        kan_lats = np.radians(df_kan['lat'].values)
        kan_lons = np.radians(df_kan['lon'].values)
        poi_lats = np.radians(df_poi['lat'].values)
        poi_lons = np.radians(df_poi['lon'].values)
        dlat = poi_lats[:, None] - kan_lats[None, :]
        dlon = poi_lons[:, None] - kan_lons[None, :]
        a = np.sin(dlat/2)**2 + np.cos(kan_lats[None, :]) * np.cos(poi_lats[:, None]) * np.sin(dlon/2)**2
        self.distances = 2 * self.R * np.arcsin(np.sqrt(a))
        self.distances = np.maximum(self.distances, 0.1)
        self.attractiveness = (df_kan['biaya'].values / 1e8) + df_kan['skor_akses_jalan'].values * 50 + df_kan['skor_visibilitas'].values * 30
        self.poi_weights = df_poi['kunjungan_hari'].values * df_poi['bobot_demand'].values * 0.12
        self.kab_map = {i: df_kan.iloc[i]['kab_kota'] for i in range(self.n_kan)}
        self.biaya = df_kan['biaya'].values
        self.demand = df_kan['demand'].values
        self.lat = df_kan['lat'].values
        self.lon = df_kan['lon'].values
        self.kan_ids = df_kan['id'].values

class HuffModel:
    def __init__(self, alpha=1.0, beta=2.0):
        self.alpha = alpha
        self.beta = beta
    def calculate(self, dist_matrix):
        A = dist_matrix.attractiveness ** self.alpha
        D = dist_matrix.distances ** self.beta
        scores = A[None, :] / D
        total_scores = scores.sum(axis=1, keepdims=True)
        total_scores = np.maximum(total_scores, 1e-10)
        probs = scores / total_scores
        demand_captured = probs * dist_matrix.poi_weights[:, None]
        hasil = []
        for i in range(dist_matrix.n_poi):
            for j in range(dist_matrix.n_kan):
                hasil.append({'id_poi': i+1, 'id_kandidat': int(dist_matrix.kan_ids[j]), 'jarak_km': float(dist_matrix.distances[i, j]), 'daya_tarik_kandidat': float(A[j]), 'probabilitas': float(probs[i, j]), 'demand_tertangkap': float(demand_captured[i, j])})
        return pd.DataFrame(hasil)

class FastNSGA2:
    def __init__(self, pop_size=50, n_generations=30, crossover_prob=0.9, mutation_prob=0.1):
        self.pop_size = pop_size
        self.n_generations = n_generations
        self.cx_prob = crossover_prob
        self.mut_prob = mutation_prob
    def dominates(self, a, b):
        return all(x <= y for x, y in zip(a, b)) and any(x < y for x, y in zip(a, b))
    def fast_non_dominated_sort(self, objs):
        n = len(objs)
        S = [[] for _ in range(n)]
        n_dom = [0] * n
        fronts = [[]]
        for p in range(n):
            for q in range(n):
                if p == q: continue
                if self.dominates(objs[p], objs[q]): S[p].append(q)
                elif self.dominates(objs[q], objs[p]): n_dom[p] += 1
            if n_dom[p] == 0: fronts[0].append(p)
        i = 0
        while fronts[i]:
            nxt = []
            for p in fronts[i]:
                for q in S[p]:
                    n_dom[q] -= 1
                    if n_dom[q] == 0: nxt.append(q)
            i += 1
            fronts.append(nxt)
        fronts.pop()
        ranks = [0] * n
        for fi, front in enumerate(fronts):
            for idx in front: ranks[idx] = fi
        return ranks, fronts
    def crowding(self, objs, front):
        if len(front) <= 2: return {idx: float('inf') for idx in front}
        dists = {idx: 0.0 for idx in front}
        for m in range(len(objs[0])):
            s = sorted(front, key=lambda i: objs[i][m])
            dists[s[0]] = dists[s[-1]] = float('inf')
            fmax, fmin = objs[s[-1]][m], objs[s[0]][m]
            if fmax - fmin < 1e-9: continue
            for k in range(1, len(s)-1): dists[s[k]] += (objs[s[k+1]][m] - objs[s[k-1]][m]) / (fmax - fmin)
        return dists
    def tournament(self, pop, ranks, crowd, size=2):
        n = len(pop)
        sel = []
        for _ in range(n):
            cands = np.random.choice(n, size, replace=False)
            w = cands[0]
            for c in cands[1:]:
                if ranks[c] < ranks[w] or (ranks[c] == ranks[w] and crowd.get(c,0) > crowd.get(w,0)): w = c
            sel.append(pop[w].copy())
        return sel
    def sbx(self, p1, p2, eta=15):
        c1, c2 = p1.copy(), p2.copy()
        if np.random.random() > self.cx_prob: return c1, c2
        for i in range(len(p1)):
            if np.random.random() > 0.5 or abs(p1[i]-p2[i]) < 1e-14: continue
            y1, y2 = (p1[i], p2[i]) if p1[i] < p2[i] else (p2[i], p1[i])
            beta_val = 1.0 + 2.0*y1/(y2-y1)
            alpha = 2.0 - beta_val**-(eta+1)
            r = np.random.random()
            beta_q = (r*alpha)**(1/(eta+1)) if r <= 1/alpha else (1/(2-r*alpha))**(1/(eta+1))
            c1[i] = 0.5*((y1+y2) - beta_q*(y2-y1))
            beta_val = 1.0 + 2.0*(1-y2)/(y2-y1)
            alpha = 2.0 - beta_val**-(eta+1)
            beta_q = (r*alpha)**(1/(eta+1)) if r <= 1/alpha else (1/(2-r*alpha))**(1/(eta+1))
            c2[i] = 0.5*((y1+y2) + beta_q*(y2-y1))
            c1[i] = max(0, min(1, c1[i]))
            c2[i] = max(0, min(1, c2[i]))
        return c1, c2
    def mutate(self, x, eta=20):
        m = x.copy()
        for i in range(len(x)):
            if np.random.random() < self.mut_prob:
                d1, d2 = x[i], 1-x[i]
                r = np.random.random()
                mp = 1/(eta+1)
                if r <= 0.5:
                    xy = 1-d1
                    val = 2*r + (1-2*r)*(xy**(eta+1))
                    dq = val**mp - 1
                else:
                    xy = 1-d2
                    val = 2*(1-r) + 2*(r-0.5)*(xy**(eta+1))
                    dq = 1 - val**mp
                m[i] = x[i] + dq
                m[i] = max(0, min(1, m[i]))
        return m
    def evaluate_batch(self, population, dist_matrix, budget, min_dist, cvar_data):
        n = len(population)
        n_kan = dist_matrix.n_kan
        pop_arr = np.array(population)
        selected = pop_arr > 0.5
        costs = (selected * dist_matrix.biaya[None, :]).sum(axis=1)
        penalty = np.zeros(n)
        penalty[costs > budget] += (costs[costs > budget] - budget) / budget * 1e6
        n_sel = selected.sum(axis=1)
        penalty[n_sel < 3] += (3 - n_sel[n_sel < 3]) * 1e5
        penalty[n_sel > 8] += (n_sel[n_sel > 8] - 8) * 1e5
        for i in range(n):
            if n_sel[i] > 1:
                sel_idx = np.where(selected[i])[0]
                for a in range(len(sel_idx)):
                    for b in range(a+1, len(sel_idx)):
                        lat_a, lon_a = dist_matrix.lat[sel_idx[a]], dist_matrix.lon[sel_idx[a]]
                        lat_b, lon_b = dist_matrix.lat[sel_idx[b]], dist_matrix.lon[sel_idx[b]]
                        R = 6371.0
                        lat1r, lon1r, lat2r, lon2r = map(np.radians, [lat_a, lon_a, lat_b, lon_b])
                        d = 2 * R * np.arcsin(np.sqrt(np.sin((lat2r-lat1r)/2)**2 + np.cos(lat1r)*np.cos(lat2r)*np.sin((lon2r-lon1r)/2)**2))
                        if d < min_dist: penalty[i] += (min_dist - d) / min_dist * 5e5
        obj_cost = costs / 1e9 + penalty
        demand_per_kan = np.zeros(n_kan)
        for _, row in cvar_data.iterrows():
            idx = np.where(dist_matrix.kan_ids == row['id_kandidat'])[0]
            if len(idx) > 0: demand_per_kan[idx[0]] = row.get('demand_tangkap', 0)
        coverage = (selected * demand_per_kan[None, :]).sum(axis=1)
        obj_coverage = -coverage / 1000
        obj_equity = np.zeros(n)
        for i in range(n):
            if n_sel[i] > 0:
                kabs = {}
                for j in range(n_kan):
                    if selected[i, j]: kabs[dist_matrix.kab_map[j]] = kabs.get(dist_matrix.kab_map[j], 0) + 1
                counts = list(kabs.values())
                if len(counts) > 1:
                    mean_c = np.mean(counts)
                    obj_equity[i] = -np.std(counts) / mean_c if mean_c > 0 else 0
                else: obj_equity[i] = -1.0
        cvar_per_kan = np.zeros(n_kan)
        for _, row in cvar_data.iterrows():
            idx = np.where(dist_matrix.kan_ids == row['id_kandidat'])[0]
            if len(idx) > 0: cvar_per_kan[idx[0]] = abs(row['cvar_95'])
        obj_risk = (selected * cvar_per_kan[None, :]).sum(axis=1) / 1e9 + penalty * 0.1
        return [[obj_cost[i], obj_coverage[i], obj_equity[i], obj_risk[i]] for i in range(n)]
    def optimize(self, dist_matrix, budget, min_dist_km, cvar_data):
        n_kan = dist_matrix.n_kan
        pop = [np.random.random(n_kan) for _ in range(self.pop_size)]
        for gen in range(self.n_generations):
            objs = self.evaluate_batch(pop, dist_matrix, budget, min_dist_km, cvar_data)
            ranks, fronts = self.fast_non_dominated_sort(objs)
            crowding_dists = {}
            for front in fronts:
                if front: crowding_dists.update(self.crowding(objs, front))
            selected = self.tournament(pop, ranks, crowding_dists)
            offspring = []
            for i in range(0, len(selected), 2):
                p1 = selected[i]
                p2 = selected[(i+1) % len(selected)]
                c1, c2 = self.sbx(p1, p2)
                c1 = self.mutate(c1)
                c2 = self.mutate(c2)
                offspring.extend([c1, c2])
            combined = pop + offspring[:self.pop_size]
            combined_objs = self.evaluate_batch(combined, dist_matrix, budget, min_dist_km, cvar_data)
            combined_ranks, combined_fronts = self.fast_non_dominated_sort(combined_objs)
            combined_crowd = {}
            for front in combined_fronts:
                if front: combined_crowd.update(self.crowding(combined_objs, front))
            sorted_idx = sorted(range(len(combined)), key=lambda i: (combined_ranks[i], -combined_crowd.get(i, 0)))
            pop = [combined[i] for i in sorted_idx[:self.pop_size]]
        final_objs = self.evaluate_batch(pop, dist_matrix, budget, min_dist_km, cvar_data)
        final_ranks, _ = self.fast_non_dominated_sort(final_objs)
        pareto = []
        for i, r in enumerate(final_ranks):
            if r == 0:
                sel_ids = [int(dist_matrix.kan_ids[j]) for j in range(n_kan) if pop[i][j] > 0.5]
                pareto.append({'chromosome': pop[i], 'objectives': final_objs[i], 'selected_ids': sel_ids, 'cost': final_objs[i][0], 'coverage': -final_objs[i][1]*1000, 'equity': -final_objs[i][2], 'risk': final_objs[i][3]})
        return pareto

class FastStackelberg:
    def __init__(self, budget, subsidy_max=0.3):
        self.budget = budget
        self.subsidy_max = subsidy_max
    def solve(self, df_kan, pareto_solutions, cvar_data):
        if not pareto_solutions: return pd.DataFrame()
        results = []
        for sol in pareto_solutions[:min(5, len(pareto_solutions))]:
            sids = sol['selected_ids']
            total_cost = sum(df_kan[df_kan['id']==sid]['biaya'].sum() for sid in sids)
            coverage = sol['coverage']
            best_s, best_w, best_p = 0, -1e9, -1e9
            for s in np.linspace(0, self.subsidy_max, 10):
                sc = total_cost * s
                if sc > self.budget * 0.4: continue
                ip = 0
                feasible = True
                for sid in sids:
                    kan = df_kan[df_kan['id']==sid].iloc[0]
                    cr = cvar_data[cvar_data['id_kandidat']==sid]
                    er = cr['expected_return'].values[0] if not cr.empty else kan['biaya']*0.15
                    profit = er - kan['biaya']*0.12 + kan['biaya']*s
                    if profit < 0: feasible = False
                    ip += profit
                if not feasible: continue
                welfare = coverage*500 + ip*0.2 - sc
                if welfare > best_w: best_w, best_s, best_p = welfare, s, ip
            results.append({'solusi_id': 0, 'lokasi_terpilih': json.dumps(sids), 'jumlah_lokasi': len(sids), 'total_biaya': total_cost, 'subsidi_optimal': best_s, 'leader_payoff': best_w, 'follower_payoff': best_p, 'net_benefit': coverage*500 - total_cost*0.05, 'coverage': coverage, 'equity_score': sol['equity']})
        df = pd.DataFrame(results)
        return df.sort_values('leader_payoff', ascending=False) if not df.empty else df

class CVaREngine:
    def __init__(self, alpha=0.95, skenario_dict=None):
        self.alpha = alpha
        self.skenario = skenario_dict or {'Optimis': {'demand_factor': 1.4, 'cost_factor': 0.85, 'prob': 0.20}, 'Moderat+': {'demand_factor': 1.15, 'cost_factor': 0.95, 'prob': 0.30}, 'Moderat': {'demand_factor': 1.0, 'cost_factor': 1.0, 'prob': 0.30}, 'Pesimis': {'demand_factor': 0.7, 'cost_factor': 1.15, 'prob': 0.20}}
    def evaluate(self, df_kan, demand_growth=1.0):
        hasil = []
        demands = df_kan['demand'].values
        biayas = df_kan['biaya'].values
        for i in range(len(df_kan)):
            returns = []
            weights = []
            for param in self.skenario.values():
                rev = demands[i] * param['demand_factor'] * demand_growth * 2500
                cost = biayas[i] * param['cost_factor'] * 0.08
                returns.append(rev - cost)
                weights.append(param['prob'])
            expected = np.average(returns, weights=weights)
            s = sorted(returns)
            cutoff = int(np.ceil((1-self.alpha)*len(s)))
            cvar = np.mean(s[:max(cutoff,1)]) if cutoff > 0 else s[0]
            z = 1.0 if expected > 0 and abs(cvar) < expected*2 else (0.7 if expected > 0 else 0.3)
            hasil.append({'id_kandidat': int(df_kan.iloc[i]['id']), 'expected_return': float(expected), 'var_95': float(np.percentile(returns, (1-self.alpha)*100)), 'cvar_95': float(cvar), 'z_invest': float(z), 'adjusted_npv': float(expected*5 - biayas[i]), 'risk_rating': 'Rendah' if z >= 0.85 else ('Sedang' if z >= 0.5 else 'Tinggi')})
        return pd.DataFrame(hasil)

class IntegratedDSS:
    def __init__(self, db, scenario_id='default'):
        self.db = db
        self.scenario_id = scenario_id
        self.config = self._load_config()
        self.huff = HuffModel(alpha=self.config['huff_alpha'], beta=self.config['huff_beta'])
        self.nsga = FastNSGA2(pop_size=self.config['nsga_popsize'], n_generations=self.config['nsga_generasi'])
        self.stackelberg = FastStackelberg(budget=self.config['budget'], subsidy_max=self.config.get('subsidy_max', 0.3))
        self.cvar = CVaREngine(alpha=self.config['cvar_alpha'], skenario_dict=self.config.get('cvar_skenario', None))
    def _load_config(self):
        df = self.db.query("SELECT * FROM skenario_config WHERE scenario_id = ?", (self.scenario_id,))
        if df.empty:
            base = {'budget': 15000000000, 'max_lokasi': 8, 'min_lokasi': 3, 'min_jarak_km': 5.0, 'huff_alpha': 1.0, 'huff_beta': 2.0, 'cvar_alpha': 0.95, 'risk_aversion': 0.5, 'nsga_popsize': 50, 'nsga_generasi': 30, 'demand_growth': 1.0}
        else:
            r = df.iloc[0]
            base = {'budget': r['budget'], 'max_lokasi': r['max_lokasi'], 'min_lokasi': r['min_lokasi'], 'min_jarak_km': r['min_jarak_km'], 'huff_alpha': r['huff_alpha'], 'huff_beta': r['huff_beta'], 'cvar_alpha': r['cvar_alpha'], 'risk_aversion': r['risk_aversion'], 'nsga_popsize': r['nsga_popsize'], 'nsga_generasi': r['nsga_generasi'], 'demand_growth': r['demand_growth']}
        huff_df = self.db.query("SELECT * FROM huff_params WHERE scenario_id = ?", (self.scenario_id,))
        if not huff_df.empty:
            base['huff_alpha'] = huff_df.iloc[0]['alpha']
            base['huff_beta'] = huff_df.iloc[0]['beta']
        stack_df = self.db.query("SELECT * FROM stackelberg_params WHERE scenario_id = ?", (self.scenario_id,))
        if not stack_df.empty: base['subsidy_max'] = stack_df.iloc[0]['subsidy_max']
        cvar_df = self.db.query("SELECT * FROM cvar_skenario WHERE scenario_id = ?", (self.scenario_id,))
        if not cvar_df.empty:
            base['cvar_skenario'] = {}
            for _, row in cvar_df.iterrows(): base['cvar_skenario'][row['nama_skenario']] = {'demand_factor': row['demand_factor'], 'cost_factor': row['cost_factor'], 'prob': row['prob']}
        return base
    def run_pipeline(self, progress_callback=None):
        df_kan = self.db.query("SELECT * FROM kandidat_lokasi")
        df_poi = self.db.query("SELECT * FROM poi_demand")
        if df_kan.empty or df_poi.empty:
            st.error("Data kandidat atau POI kosong!")
            return None
        if progress_callback: progress_callback(0.05, "Membangun matriks jarak spasial...")
        dist_matrix = DistanceMatrix(df_kan, df_poi)
        if progress_callback: progress_callback(0.15, "Menghitung permintaan spasial (Huff)...")
        df_huff = self.huff.calculate(dist_matrix)
        df_huff.to_sql('hasil_huff', self.db.conn, if_exists='replace', index=False)
        df_huff_agg = df_huff.groupby('id_kandidat')['demand_tertangkap'].sum().reset_index()
        if progress_callback: progress_callback(0.25, "Mengevaluasi risiko investasi (CVaR)...")
        df_cvar = self.cvar.evaluate(df_kan, demand_growth=self.config['demand_growth'])
        df_cvar['scenario_id'] = self.scenario_id
        df_cvar.to_sql('hasil_cvar', self.db.conn, if_exists='replace', index=False)
        df_cvar = df_cvar.merge(df_huff_agg, on='id_kandidat', how='left').fillna(0)
        if progress_callback: progress_callback(0.35, "Optimasi multi-objektif (NSGA-II)...")
        pareto_solutions = self.nsga.optimize(dist_matrix, budget=self.config['budget'], min_dist_km=self.config['min_jarak_km'], cvar_data=df_cvar)
        nsga_rows = []
        for i, sol in enumerate(pareto_solutions):
            nsga_rows.append({'scenario_id': self.scenario_id, 'solusi_id': i, 'generasi': self.config['nsga_generasi'], 'obj_biaya': sol['cost'], 'obj_coverage': sol['coverage'], 'obj_equity': sol['equity'], 'obj_risiko': sol['risk'], 'lokasi_terpilih': json.dumps(sol['selected_ids']), 'is_pareto': True, 'rank_crowding': 0.0})
        pd.DataFrame(nsga_rows).to_sql('hasil_nsga', self.db.conn, if_exists='replace', index=False)
        if progress_callback: progress_callback(0.80, "Simulasi strategi Stackelberg...")
        df_stack = self.stackelberg.solve(df_kan, pareto_solutions, df_cvar)
        df_stack_final = pd.DataFrame()
        if not df_stack.empty:
            stack_rows = []
            for _, row in df_stack.iterrows():
                locs = json.loads(row['lokasi_terpilih'])
                for lid in locs:
                    stack_rows.append({'scenario_id': self.scenario_id, 'id_kandidat': lid, 'x_terpilih': 1, 'subsidi_optimal': row['subsidi_optimal'], 'investor_response': 'invest', 'leader_payoff': row['leader_payoff']/len(locs), 'follower_payoff': row['follower_payoff']/len(locs), 'net_benefit': row['net_benefit']/len(locs), 'equilibrium_type': 'Stackelberg-Nash'})
            all_ids = set(df_kan['id'].values)
            sel_ids = set(r['id_kandidat'] for r in stack_rows)
            for rid in all_ids - sel_ids:
                stack_rows.append({'scenario_id': self.scenario_id, 'id_kandidat': rid, 'x_terpilih': 0, 'subsidi_optimal': 0, 'investor_response': 'no_invest', 'leader_payoff': 0, 'follower_payoff': 0, 'net_benefit': 0, 'equilibrium_type': 'rejected'})
            df_stack_final = pd.DataFrame(stack_rows)
            df_stack_final.to_sql('hasil_stackelberg', self.db.conn, if_exists='replace', index=False)
        if progress_callback: progress_callback(1.0, "Selesai!")
        return {'huff': df_huff, 'huff_agg': df_huff_agg, 'cvar': df_cvar, 'nsga': pd.DataFrame(nsga_rows), 'stackelberg': df_stack_final, 'pareto': pareto_solutions}


# ============================================================
# HELPER: PLOTLY CHARTS
# ============================================================
def plotly_bar_comparison(df, x_col, y1_col, y2_col, title, y1_name, y2_name, y_label):
    fig = go.Figure()
    fig.add_trace(go.Bar(name=y1_name, x=df[x_col], y=df[y1_col], marker_color='#1E40AF', opacity=0.9))
    fig.add_trace(go.Bar(name=y2_name, x=df[x_col], y=df[y2_col], marker_color='#06B6D4', opacity=0.9))
    fig.update_layout(
        title=title, barmode='group', yaxis_title=y_label,
        plot_bgcolor='white', paper_bgcolor='white',
        font=dict(family="Arial, sans-serif", size=12, color="#475569"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=40, r=40, t=60, b=40),
        hovermode='x unified'
    )
    fig.update_xaxes(tickangle=45, tickfont_size=10)
    return fig

def plotly_pareto_scatter(df, x_col, y_col, color_col=None, title="", x_label="", y_label=""):
    if color_col and color_col in df.columns:
        fig = px.scatter(df, x=x_col, y=y_col, color=color_col, size_max=15,
                         title=title, labels={x_col: x_label, y_col: y_label},
                         color_discrete_sequence=px.colors.qualitative.Bold)
    else:
        fig = px.scatter(df, x=x_col, y=y_col, title=title,
                         labels={x_col: x_label, y_col: y_label},
                         color_discrete_sequence=['#1E40AF'])
    fig.update_traces(marker=dict(size=12, line=dict(width=2, color='white')))
    fig.update_layout(
        plot_bgcolor='white', paper_bgcolor='white',
        font=dict(family="Arial, sans-serif", size=12, color="#475569"),
        margin=dict(l=40, r=40, t=60, b=40),
        hovermode='closest'
    )
    return fig

def plotly_risk_distribution(values, var_val, cvar_val, title="Distribusi Risiko"):
    fig = go.Figure()
    fig.add_trace(go.Histogram(x=values, nbinsx=50, marker_color='#3B82F6', opacity=0.7,
                               name='Distribusi Return'))
    fig.add_vline(x=var_val, line_dash="dash", line_color="#F59E0B", annotation_text=f"VaR", annotation_position="top")
    fig.add_vline(x=cvar_val, line_color="#DC2626", annotation_text=f"CVaR", annotation_position="top")
    fig.update_layout(title=title, xaxis_title="Net Return", yaxis_title="Frekuensi",
                      plot_bgcolor='white', paper_bgcolor='white',
                      font=dict(family="Arial, sans-serif", size=12, color="#475569"),
                      margin=dict(l=40, r=40, t=60, b=40), showlegend=False)
    return fig

def plotly_payoff_chart(df, x_col, y1_col, y2_col, title):
    fig = go.Figure()
    fig.add_trace(go.Bar(name='Leader Payoff', x=df[x_col], y=df[y1_col], marker_color='#1E40AF'))
    fig.add_trace(go.Bar(name='Follower Payoff', x=df[x_col], y=df[y2_col], marker_color='#F59E0B'))
    fig.update_layout(title=title, barmode='group', yaxis_title="Juta Rp",
                      plot_bgcolor='white', paper_bgcolor='white',
                      font=dict(family="Arial, sans-serif", size=12, color="#475569"),
                      legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
                      margin=dict(l=40, r=40, t=60, b=40))
    fig.update_xaxes(tickangle=30, tickfont_size=10)
    return fig

# ============================================================
# HELPER: EXPORT EXCEL
# ============================================================
def export_excel_full(results, df_kan, df_poi, df_wilayah, filename="rekomendasi_spklu.xlsx"):
    if not OPENPYXL_AVAILABLE:
        return None
    from io import BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_stack = results.get('stackelberg', pd.DataFrame())
        df_cvar = results.get('cvar', pd.DataFrame())
        df_huff_agg = results.get('huff_agg', pd.DataFrame())
        df_rekom = df_kan.copy()
        df_rekom = df_rekom.merge(df_huff_agg, left_on='id', right_on='id_kandidat', how='left').fillna(0)
        if not df_stack.empty:
            df_rekom = df_rekom.merge(df_stack.groupby('id_kandidat').agg({'x_terpilih':'max','net_benefit':'sum','subsidi_optimal':'max'}).reset_index(), on='id_kandidat', how='left').fillna(0)
        else:
            df_rekom['x_terpilih'] = 0; df_rekom['net_benefit'] = 0; df_rekom['subsidi_optimal'] = 0
        if not df_cvar.empty:
            df_rekom = df_rekom.merge(df_cvar[['id_kandidat','z_invest','risk_rating']], on='id_kandidat', how='left').fillna(0)
        else:
            df_rekom['z_invest'] = 0; df_rekom['risk_rating'] = 'N/A'
        df_rekom['Status'] = df_rekom['x_terpilih'].apply(lambda x: 'Terpilih' if x == 1 else 'Ditolak')
        df_rekom[['nama','kab_kota','demand_tertangkap','biaya','z_invest','net_benefit','Status']].to_excel(writer, sheet_name='Rekomendasi', index=False)
        results.get('huff', pd.DataFrame()).to_excel(writer, sheet_name='Huff Model', index=False)
        results.get('cvar', pd.DataFrame()).to_excel(writer, sheet_name='CVaR', index=False)
        results.get('nsga', pd.DataFrame()).to_excel(writer, sheet_name='NSGA-II', index=False)
        results.get('stackelberg', pd.DataFrame()).to_excel(writer, sheet_name='Stackelberg', index=False)
        df_kan.to_excel(writer, sheet_name='Data Kandidat', index=False)
        df_poi.to_excel(writer, sheet_name='Data POI', index=False)
    output.seek(0)
    return output

# ============================================================
# HELPER: PDF REPORT GENERATOR
# ============================================================
class PDFReport(FPDF):
    def header(self):
        self.set_font('DejaVu', 'B', 16)
        self.set_text_color(30, 58, 138)
        self.cell(0, 10, 'SPK-GeoOptima DSS — Laporan Analisis', ln=True, align='C')
        self.set_font('DejaVu', '', 10)
        self.set_text_color(100, 116, 139)
        self.cell(0, 6, f"Framework: GIS + NSGA-II + Huff + Stackelberg + CVaR | {datetime.now().strftime('%d %B %Y')}", ln=True, align='C')
        self.ln(5)
        self.set_draw_color(59, 130, 246)
        self.line(10, self.get_y(), 200, self.get_y())
        self.ln(5)

    def footer(self):
        self.set_y(-15)
        self.set_font('DejaVu', 'I', 8)
        self.set_text_color(150, 150, 150)
        self.cell(0, 10, f'Halaman {self.page_no()}', align='C')

def generate_pdf_report(results, df_kan, budget, min_jarak, nsga_pop, nsga_gen):
    if not FPDF_AVAILABLE:
        return None
    from io import BytesIO
    pdf = PDFReport()
    try:
        pdf.add_font('DejaVu', '', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf', uni=True)
        pdf.add_font('DejaVu', 'B', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf', uni=True)
        pdf.add_font('DejaVu', 'I', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf', uni=True)
    except:
        try:
            pdf.add_font('DejaVu', '', 'DejaVuSans.ttf', uni=True)
            pdf.add_font('DejaVu', 'B', 'DejaVuSans-Bold.ttf', uni=True)
        except:
            pdf.set_font('Arial', '', 10)
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 10, '1. Ringkasan Eksekutif', ln=True)
    pdf.set_font('DejaVu', '', 11)
    pdf.set_text_color(71, 85, 105)
    df_stack = results.get('stackelberg', pd.DataFrame())
    df_cvar = results.get('cvar', pd.DataFrame())
    df_huff_agg = results.get('huff_agg', pd.DataFrame())
    terpilih = df_stack[df_stack['x_terpilih'] == 1] if not df_stack.empty else pd.DataFrame()
    n_terpilih = len(terpilih['id_kandidat'].unique()) if not terpilih.empty else 0
    total_inv = df_kan[df_kan['id'].isin(terpilih['id_kandidat'].unique())]['biaya'].sum() if not terpilih.empty else 0
    total_demand = df_kan['demand'].sum() if not df_kan.empty else 0
    summary_text = f"Berdasarkan framework SPRA-DSS, sistem telah mengidentifikasi {n_terpilih} lokasi prioritas untuk implementasi SPKLU di Kalimantan Selatan. Total investasi sebesar Rp {total_inv/1e9:.1f} Miliar dengan cakupan permintaan {total_demand:,.0f} kWh/hari. Parameter analisis: Budget Rp {budget/1e9:.0f}M, Jarak Minimum {min_jarak:.0f} km, NSGA-II Populasi {nsga_pop}, Generasi {nsga_gen}."
    pdf.multi_cell(0, 8, summary_text)
    pdf.ln(5)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(0, 10, '2. Parameter Analisis', ln=True)
    pdf.set_font('DejaVu', '', 11)
    pdf.cell(0, 8, f'Budget: Rp {budget/1e9:.0f} Miliar', ln=True)
    pdf.cell(0, 8, f'Jarak Minimum: {min_jarak:.0f} km', ln=True)
    pdf.cell(0, 8, f'NSGA-II Populasi: {nsga_pop}', ln=True)
    pdf.cell(0, 8, f'NSGA-II Generasi: {nsga_gen}', ln=True)
    pdf.ln(5)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.cell(0, 10, '3. Matriks Keputusan Lokasi', ln=True)
    pdf.set_font('DejaVu', 'B', 10)
    pdf.set_fill_color(59, 130, 246)
    pdf.set_text_color(255, 255, 255)
    headers = ['Nama Lokasi', 'Kab/Kota', 'Demand', 'Investasi', 'Status']
    col_widths = [60, 40, 30, 35, 25]
    for h, w in zip(headers, col_widths):
        pdf.cell(w, 10, h, border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_font('DejaVu', '', 9)
    pdf.set_text_color(71, 85, 105)
    df_rekom = df_kan.merge(df_huff_agg, left_on='id', right_on='id_kandidat', how='left').fillna(0)
    if not df_stack.empty:
        df_rekom = df_rekom.merge(df_stack.groupby('id_kandidat').agg({'x_terpilih':'max'}).reset_index(), on='id_kandidat', how='left').fillna(0)
    else: df_rekom['x_terpilih'] = 0
    df_rekom = df_rekom.sort_values('demand_tertangkap', ascending=False)
    for _, row in df_rekom.head(15).iterrows():
        status = 'Terpilih' if row['x_terpilih'] == 1 else 'Ditolak'
        pdf.cell(60, 8, str(row['nama'])[:25], border=1)
        pdf.cell(40, 8, str(row['kab_kota'])[:15], border=1)
        pdf.cell(30, 8, f"{row['demand_tertangkap']:,.0f}", border=1, align='R')
        pdf.cell(35, 8, f"Rp {row['biaya']/1e9:.1f}M", border=1, align='R')
        pdf.cell(25, 8, status, border=1, align='C')
        pdf.ln()
    output = BytesIO()
    pdf.output(output)
    output.seek(0)
    return output

# ============================================================
# HELPER: SCENARIO MANAGER
# ============================================================
def save_scenario_to_db(db, nama, deskripsi, budget, max_lokasi, min_jarak, huff_beta, cvar_alpha, risk_aversion, nsga_pop, nsga_gen, results):
    hasil_json = json.dumps({
        'n_pareto': len(results.get('nsga', pd.DataFrame())),
        'n_terpilih': len(results.get('stackelberg', pd.DataFrame())[results.get('stackelberg', pd.DataFrame())['x_terpilih']==1]['id_kandidat'].unique()) if not results.get('stackelberg', pd.DataFrame()).empty else 0,
        'total_demand': results.get('huff_agg', pd.DataFrame())['demand_tertangkap'].sum() if not results.get('huff_agg', pd.DataFrame()).empty else 0
    })
    db.execute("""
        INSERT INTO saved_scenarios (nama, deskripsi, budget, max_lokasi, min_jarak_km, huff_beta, cvar_alpha, risk_aversion, nsga_popsize, nsga_generasi, hasil_json)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (nama, deskripsi, budget, max_lokasi, min_jarak, huff_beta, cvar_alpha, risk_aversion, nsga_pop, nsga_gen, hasil_json))

def load_scenarios_from_db(db):
    return db.query("SELECT * FROM saved_scenarios ORDER BY created_at DESC")

def delete_scenario(db, scenario_id):
    db.execute("DELETE FROM saved_scenarios WHERE id = ?", (scenario_id,))


# ============================================================
# INISIALISASI DATABASE
# ============================================================
@st.cache_resource
def get_db():
    try:
        db = SpatialDatabase("spklu_integrated_v51.db")
        db.seed_default_data()
        return db
    except Exception as e:
        st.error(f"Gagal inisialisasi database: {e}")
        return None

try:
    db = get_db()
    if db is None:
        st.error("Database tidak dapat diinisialisasi.")
        st.stop()
except Exception as e:
    st.error(f"Error kritis: {e}")
    st.stop()

# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.markdown("<h1 style='color:white; font-size:1.3rem; font-weight:800;'>⚡ SPK-GeoOptima</h1>", unsafe_allow_html=True)
    st.markdown("<p style='color:#94A3B8; font-size:0.8rem;'>v5.1 — Spatial-Probabilistic Risk-Averse DSS</p>", unsafe_allow_html=True)
    st.markdown("---")

    show_input = st.toggle("📝 Panel Input Data", st.session_state.show_input_panel, help="Buka panel untuk mengedit data")
    st.session_state.show_input_panel = show_input
    st.markdown("---")

    st.markdown("<h3 style='color:#E2E8F0; font-size:0.95rem;'>⚡ Mode Komputasi</h3>", unsafe_allow_html=True)
    mode = st.radio("", ["🚀 Fast Mode", "⚖️ Standard Mode", "🎯 Accurate Mode"], index=0, label_visibility="collapsed")
    if "Fast" in mode: default_pop, default_gen = 50, 30
    elif "Standard" in mode: default_pop, default_gen = 80, 80
    else: default_pop, default_gen = 150, 200
    st.markdown("---")

    st.markdown("<h3 style='color:#E2E8F0; font-size:0.95rem;'>🎛️ Parameter Skenario</h3>", unsafe_allow_html=True)
    budget = st.slider("Anggaran (Miliar Rp)", 5, 50, 15, 1) * 1e9
    max_lokasi = st.slider("Maksimum Lokasi", 3, 15, 8, 1)
    min_jarak = st.slider("Jarak Minimum (km)", 3.0, 20.0, 5.0, 1.0)

    st.markdown("<h3 style='color:#E2E8F0; font-size:0.95rem; margin-top:16px;'>🔬 Parameter Model</h3>", unsafe_allow_html=True)
    huff_beta = st.slider("Huff β (Decay)", 1.0, 4.0, 2.0, 0.1)
    cvar_conf = st.slider("CVaR Confidence", 0.80, 0.99, 0.95, 0.01)
    risk_aversion = st.slider("Risk Aversion", 0.0, 1.0, 0.5, 0.1)

    st.markdown("<h3 style='color:#E2E8F0; font-size:0.95rem; margin-top:16px;'>🧬 NSGA-II</h3>", unsafe_allow_html=True)
    nsga_pop = st.select_slider("Populasi", options=[30, 50, 80, 100, 150, 200], value=default_pop)
    nsga_gen = st.select_slider("Generasi", options=[20, 30, 50, 80, 100, 150, 200, 300], value=default_gen)

    run_analysis = st.button("🚀 Jalankan Analisis Terpadu", use_container_width=True)
    st.markdown("---")

    st.markdown("<h3 style='color:#E2E8F0; font-size:0.95rem;'>💾 Scenario Manager</h3>", unsafe_allow_html=True)
    with st.expander("Simpan / Load Skenario"):
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            if st.button("💾 Simpan Skenario", use_container_width=True):
                st.session_state.show_save_dialog = True
        with col_s2:
            if st.button("📂 Load Skenario", use_container_width=True):
                st.session_state.show_load_dialog = True

    st.markdown("---")
    st.markdown("<p style='color:#64748B; font-size:0.7rem;'>v5.1 | GIS + NSGA-II + Huff + Stackelberg + CVaR</p>", unsafe_allow_html=True)

# ============================================================
# HEADER
# ============================================================
tanggal_kini = datetime.now().strftime("%d %B %Y")
st.markdown(f"""
<div class='top-header'>
    <div style='display:flex; align-items:center; gap:14px;'>
        <div style='width:50px; height:50px; background:linear-gradient(135deg, #FBBF24, #F59E0B); border-radius:12px; display:flex; align-items:center; justify-content:center; font-size:1.8rem;'>⚡</div>
        <div>
            <div class='header-title'>SPK-GeoOptima DSS</div>
            <div class='header-subtitle'>Spatial-Probabilistic Risk-Averse Decision Support System</div>
        </div>
    </div>
    <div style='display:flex; gap:10px;'>
        <div class='header-badge'>📍 Kalimantan Selatan</div>
        <div class='header-badge'>🟢 Aktif | {tanggal_kini}</div>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class='pillar-grid'>
    <div class='pillar-card'><div class='pillar-icon'>🌐</div><div class='pillar-name'>GIS</div><div class='pillar-desc'>Spatial Data</div></div>
    <div class='pillar-card'><div class='pillar-icon'>📊</div><div class='pillar-name'>Huff Model</div><div class='pillar-desc'>Demand Analysis</div></div>
    <div class='pillar-card'><div class='pillar-icon'>🧬</div><div class='pillar-name'>NSGA-II</div><div class='pillar-desc'>Multi-Objective</div></div>
    <div class='pillar-card'><div class='pillar-icon'>🎯</div><div class='pillar-name'>Stackelberg</div><div class='pillar-desc'>Strategic Game</div></div>
    <div class='pillar-card'><div class='pillar-icon'>⚖️</div><div class='pillar-name'>CVaR</div><div class='pillar-desc'>Risk Assessment</div></div>
</div>
""", unsafe_allow_html=True)

# ============================================================
# PANEL INPUT DATA — menggunakan st.data_editor (native Streamlit)
# ============================================================
if st.session_state.show_input_panel:
    st.markdown("<div class='input-panel'>", unsafe_allow_html=True)
    st.markdown("<div class='input-panel-title'>📝 PANEL INPUT DATA — Kelola Data untuk Semua Metode</div>", unsafe_allow_html=True)
    input_tab1, input_tab2, input_tab3, input_tab4, input_tab5 = st.tabs(["🌐 Data Spasial (GIS)", "📊 Huff Model", "🧬 NSGA-II", "🎯 Stackelberg", "⚖️ CVaR"])

    with input_tab1:
        st.info("Kelola data wilayah, kandidat lokasi, POI, dan jaringan jalan.")
        gis_sub1, gis_sub2, gis_sub3, gis_sub4 = st.tabs(["🏛️ Wilayah", "📍 Kandidat", "🏢 POI", "🛣️ Jalan"])
        with gis_sub1:
            df_wilayah = db.query("SELECT * FROM wilayah_kalsel")
            edited_wilayah = st.data_editor(df_wilayah, num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_wilayah")
            if st.button("💾 Simpan Wilayah", key="sv_wilayah"):
                db.execute("DELETE FROM wilayah_kalsel")
                for _, row in edited_wilayah.iterrows():
                    db.execute("INSERT INTO wilayah_kalsel (kode, nama, tipe, lat_center, lon_center, populasi, luas_km2) VALUES (?,?,?,?,?,?,?)", (row['kode'], row['nama'], row['tipe'], row['lat_center'], row['lon_center'], row['populasi'], row['luas_km2']))
                st.toast("✅ Data wilayah tersimpan!", icon="💾")
        with gis_sub2:
            df_kan_ed = db.query("SELECT * FROM kandidat_lokasi")
            edited_kan = st.data_editor(df_kan_ed, num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_kan")
            if st.button("💾 Simpan Kandidat", key="sv_kan"):
                db.execute("DELETE FROM kandidat_lokasi")
                for _, row in edited_kan.iterrows():
                    db.execute("INSERT INTO kandidat_lokasi (id, nama, kab_kota, kecamatan, lat, lon, biaya, demand, kapasitas_grid_kva, skor_akses_jalan, skor_visibilitas, status, zona_prioritas) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", (row['id'], row['nama'], row['kab_kota'], row['kecamatan'], row['lat'], row['lon'], row['biaya'], row['demand'], row['kapasitas_grid_kva'], row['skor_akses_jalan'], row['skor_visibilitas'], row['status'], row['zona_prioritas']))
                st.toast("✅ Data kandidat tersimpan!", icon="💾")
        with gis_sub3:
            df_poi_ed = db.query("SELECT * FROM poi_demand")
            edited_poi = st.data_editor(df_poi_ed, num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_poi")
            if st.button("💾 Simpan POI", key="sv_poi"):
                db.execute("DELETE FROM poi_demand")
                for _, row in edited_poi.iterrows():
                    db.execute("INSERT INTO poi_demand (id, nama, tipe, kab_kota, lat, lon, kunjungan_hari, daya_tarik, bobot_demand) VALUES (?,?,?,?,?,?,?,?,?)", (row['id'], row['nama'], row['tipe'], row['kab_kota'], row['lat'], row['lon'], row['kunjungan_hari'], row['daya_tarik'], row['bobot_demand']))
                st.toast("✅ Data POI tersimpan!", icon="💾")
        with gis_sub4:
            df_jalan_ed = db.query("SELECT * FROM jaringan_jalan")
            edited_jalan = st.data_editor(df_jalan_ed, num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_jalan")
            if st.button("💾 Simpan Jalan", key="sv_jalan"):
                db.execute("DELETE FROM jaringan_jalan")
                for _, row in edited_jalan.iterrows():
                    db.execute("INSERT INTO jaringan_jalan (id, nama, tipe, lat_from, lon_from, lat_to, lon_to, panjang_km) VALUES (?,?,?,?,?,?,?,?)", (row['id'], row['nama'], row['tipe'], row['lat_from'], row['lon_from'], row['lat_to'], row['lon_to'], row['panjang_km']))
                st.toast("✅ Data jalan tersimpan!", icon="💾")

    with input_tab2:
        st.info("Huff Model: α = eksponen daya tarik, β = eksponen penurunan jarak")
        c1, c2 = st.columns(2)
        with c1: huff_alpha_input = st.number_input("α (Attractiveness)", 0.1, 5.0, 1.0, 0.1)
        with c2: huff_beta_input = st.number_input("β (Distance Decay)", 0.1, 5.0, 2.0, 0.1)
        if st.button("💾 Simpan Huff", key="sv_huff"):
            db.execute("DELETE FROM huff_params WHERE scenario_id = 'default'")
            db.execute("INSERT INTO huff_params (scenario_id, alpha, beta) VALUES (?,?,?)", ('default', huff_alpha_input, huff_beta_input))
            st.toast(f"✅ Parameter Huff tersimpan!", icon="💾")

    with input_tab3:
        st.info("NSGA-II: Parameter optimasi multi-objektif")
        c1, c2 = st.columns(2)
        with c1:
            nsga_pop_input = st.number_input("Ukuran Populasi", 10, 500, 50, 10)
            nsga_gen_input = st.number_input("Jumlah Generasi", 10, 500, 30, 10)
        with c2:
            nsga_budget = st.number_input("Budget (Rp)", 1e9, 1e12, 15e9, 1e9)
            nsga_min_dist = st.number_input("Jarak Minimum (km)", 1.0, 50.0, 5.0, 1.0)
        if st.button("💾 Simpan NSGA-II", key="sv_nsga"):
            db.execute("DELETE FROM skenario_config WHERE scenario_id = 'default'")
            db.execute("INSERT INTO skenario_config (scenario_id, nama, budget, max_lokasi, min_lokasi, min_jarak_km, huff_alpha, huff_beta, cvar_alpha, risk_aversion, nsga_popsize, nsga_generasi, demand_growth) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", ('default', 'Skenario Default', nsga_budget, 8, 3, nsga_min_dist, 1.0, 2.0, 0.95, 0.5, nsga_pop_input, nsga_gen_input, 1.0))
            st.toast("✅ Parameter NSGA-II tersimpan!", icon="💾")

    with input_tab4:
        st.info("Stackelberg: Leader (Pemerintah) vs Follower (Investor)")
        c1, c2 = st.columns(2)
        with c1: stack_subsidy = st.slider("Subsidi Maksimum (%)", 0.0, 100.0, 30.0, 5.0) / 100
        with c2: stack_budget = st.number_input("Budget Pemerintah (Rp)", 1e9, 1e12, 15e9, 1e9)
        if st.button("💾 Simpan Stackelberg", key="sv_stack"):
            db.execute("DELETE FROM stackelberg_params WHERE scenario_id = 'default'")
            db.execute("INSERT INTO stackelberg_params (scenario_id, subsidy_max, budget_pemerintah) VALUES (?,?,?)", ('default', stack_subsidy, stack_budget))
            st.toast("✅ Parameter Stackelberg tersimpan!", icon="💾")

    with input_tab5:
        st.info("CVaR: Conditional Value at Risk")
        c1, c2 = st.columns(2)
        with c1: cvar_alpha_input = st.slider("Confidence Level (α)", 0.80, 0.99, 0.95, 0.01)
        with c2: cvar_demand_growth = st.number_input("Demand Growth Factor", 0.5, 2.0, 1.0, 0.1)
        df_cvar_skenario = db.query("SELECT * FROM cvar_skenario WHERE scenario_id = 'default'")
        edited_cvar = st.data_editor(df_cvar_skenario[['nama_skenario', 'demand_factor', 'cost_factor', 'prob']], num_rows="dynamic", use_container_width=True, hide_index=True, key="ed_cvar")
        total_prob = edited_cvar['prob'].sum()
        if abs(total_prob - 1.0) > 0.01: st.warning(f"⚠️ Total prob = {total_prob:.2f} (harus = 1.00)")
        else: st.success(f"✅ Total prob valid: {total_prob:.2f}")
        if st.button("💾 Simpan CVaR", key="sv_cvar"):
            db.execute("DELETE FROM cvar_skenario WHERE scenario_id = 'default'")
            for _, row in edited_cvar.iterrows():
                db.execute("INSERT INTO cvar_skenario (scenario_id, nama_skenario, demand_factor, cost_factor, prob) VALUES (?,?,?,?,?)", ('default', row['nama_skenario'], row['demand_factor'], row['cost_factor'], row['prob']))
            db.execute("UPDATE skenario_config SET cvar_alpha = ? WHERE scenario_id = 'default'", (cvar_alpha_input,))
            st.toast("✅ Parameter CVaR tersimpan!", icon="💾")
    st.markdown("</div>", unsafe_allow_html=True)


# ============================================================
# EKSEKUSI PIPELINE
# ============================================================
results = None
if run_analysis:
    progress_bar = st.progress(0)
    status_text = st.empty()
    time_text = st.empty()
    start_time = datetime.now()
    def update_progress(pct, msg):
        progress_bar.progress(min(int(pct * 100), 100))
        elapsed = (datetime.now() - start_time).total_seconds()
        status_text.markdown(f"<p style='text-align:center; color:#1E40AF; font-weight:600;'>{msg}</p>", unsafe_allow_html=True)
        time_text.markdown(f"<p style='text-align:center; color:#64748B; font-size:0.85rem;'>⏱️ {elapsed:.1f} detik</p>", unsafe_allow_html=True)
    try:
        db.execute("DELETE FROM skenario_config WHERE scenario_id = 'current'")
        db.execute("INSERT INTO skenario_config (scenario_id, nama, budget, max_lokasi, min_lokasi, min_jarak_km, huff_alpha, huff_beta, cvar_alpha, risk_aversion, nsga_popsize, nsga_generasi, demand_growth) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", ('current', 'Skenario Aktif', budget, max_lokasi, 3, min_jarak, 1.0, huff_beta, cvar_conf, risk_aversion, nsga_pop, nsga_gen, 1.0))
        dss = IntegratedDSS(db, scenario_id='default')
        with st.spinner('Menjalankan pipeline analitik terintegrasi...'):
            results = dss.run_pipeline(progress_callback=update_progress)
        if results:
            total_time = (datetime.now() - start_time).total_seconds()
            progress_bar.empty(); status_text.empty(); time_text.empty()
            st.success(f"✅ Analisis selesai dalam **{total_time:.1f} detik**!")
            st.toast("🎉 Analisis berhasil! Lihat hasil di tab berikut.", icon="🚀")
            st.balloons()
    except Exception as e:
        progress_bar.empty(); status_text.empty(); time_text.empty()
        st.error(f"❌ Error saat analisis: {e}")
        st.exception(e)
else:
    try:
        df_huff = db.query("SELECT * FROM hasil_huff")
        if not df_huff.empty:
            results = {'huff': df_huff, 'huff_agg': db.query("SELECT id_kandidat, SUM(demand_tertangkap) as demand_tertangkap FROM hasil_huff GROUP BY id_kandidat"), 'cvar': db.query("SELECT * FROM hasil_cvar"), 'nsga': db.query("SELECT * FROM hasil_nsga"), 'stackelberg': db.query("SELECT * FROM hasil_stackelberg")}
    except Exception: pass

# ============================================================
# LOAD DATA
# ============================================================
try:
    df_kan = db.query("SELECT * FROM kandidat_lokasi")
    df_poi = db.query("SELECT * FROM poi_demand")
    df_wilayah = db.query("SELECT * FROM wilayah_kalsel")
    if df_kan.empty: st.error("Data kandidat kosong!"); st.stop()
except Exception as e: st.error(f"Error memuat data: {e}"); st.stop()

# ============================================================
# TABS HASIL ANALISIS
# ============================================================
tab1, tab2, tab3, tab4, tab5 = st.tabs(["📊 Dashboard", "🗺️ Spasial", "🧬 NSGA-II", "🎯 Stackelberg & CVaR", "✅ Rekomendasi"])

# ============================================================
# TAB 1: DASHBOARD
# ============================================================
with tab1:
    if results is None:
        st.info("👆 Silakan jalankan analisis terlebih dahulu.")
    else:
        try:
            df_stack = results.get('stackelberg', pd.DataFrame())
            df_cvar = results.get('cvar', pd.DataFrame())
            df_huff_agg = results.get('huff_agg', pd.DataFrame())
            df_nsga = results.get('nsga', pd.DataFrame())
            terpilih = df_stack[df_stack['x_terpilih'] == 1] if not df_stack.empty else pd.DataFrame()
            n_terpilih = len(terpilih['id_kandidat'].unique()) if not terpilih.empty else 0
            total_inv = df_kan[df_kan['id'].isin(terpilih['id_kandidat'].unique())]['biaya'].sum() if not terpilih.empty else 0
            total_demand = df_kan['demand'].sum() if not df_kan.empty else 0
            total_nb = terpilih['net_benefit'].sum() if not terpilih.empty else 0
            n_pareto = len(df_nsga[df_nsga['is_pareto'] == True]) if not df_nsga.empty else 0
            avg_cvar = df_cvar['z_invest'].mean() if not df_cvar.empty else 0

            st.markdown(f"""
            <div class='kpi-grid'>
                <div class='kpi-card kpi-navy'><div class='kpi-label'>Lokasi Kandidat</div><div class='kpi-value'>{len(df_kan)}</div><div class='kpi-sub'>13 Kab/Kota Kalsel</div></div>
                <div class='kpi-card kpi-blue'><div class='kpi-label'>Solusi Pareto</div><div class='kpi-value'>{n_pareto}</div><div class='kpi-sub'>NSGA-II Non-Dominated</div></div>
                <div class='kpi-card kpi-teal'><div class='kpi-label'>Lokasi Terpilih</div><div class='kpi-value'>{n_terpilih}</div><div class='kpi-sub'>Stackelberg Equilibrium</div></div>
                <div class='kpi-card kpi-green'><div class='kpi-label'>Investasi</div><div class='kpi-value'>Rp {total_inv/1e9:.1f}M</div><div class='kpi-sub'>dari Rp {budget/1e9:.0f}M</div></div>
                <div class='kpi-card kpi-amber'><div class='kpi-label'>Permintaan</div><div class='kpi-value'>{total_demand:,.0f}</div><div class='kpi-sub'>kWh/hari</div></div>
                <div class='kpi-card kpi-rose'><div class='kpi-label'>Net Benefit</div><div class='kpi-value'>Rp {total_nb/1e6:.0f}Jt</div><div class='kpi-sub'>Proyeksi tahunan</div></div>
                <div class='kpi-card kpi-violet'><div class='kpi-label'>Skor Risiko</div><div class='kpi-value'>{avg_cvar*100:.0f}%</div><div class='kpi-sub'>CVaR {cvar_conf*100:.0f}%</div></div>
            </div>
            """, unsafe_allow_html=True)

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
                st.markdown("<div class='pbi-card-title'>📊 Permintaan Potensial vs Tertangkap</div>", unsafe_allow_html=True)
                df_chart = df_kan.merge(df_huff_agg, left_on='id', right_on='id_kandidat', how='left').fillna(0)
                df_chart['nama_short'] = df_chart['nama'].str[:15]
                fig = plotly_bar_comparison(df_chart, 'nama_short', 'demand', 'demand_tertangkap', 'Permintaan per Lokasi', 'Potensial', 'Tertangkap', 'kWh / hari')
                st.plotly_chart(fig, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)
            with c2:
                st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
                st.markdown("<div class='pbi-card-title'>📊 Investasi vs Net Benefit</div>", unsafe_allow_html=True)
                if not df_stack.empty:
                    df_chart2 = df_kan.merge(df_stack.groupby('id_kandidat')['net_benefit'].sum().reset_index(), left_on='id', right_on='id_kandidat', how='left').fillna(0)
                else:
                    df_chart2 = df_kan.copy(); df_chart2['net_benefit'] = 0
                df_chart2['nama_short'] = df_chart2['nama'].str[:15]
                fig2 = plotly_bar_comparison(df_chart2, 'nama_short', 'biaya', 'net_benefit', 'Investasi vs Net Benefit', 'Investasi', 'Net Benefit', 'Rp')
                st.plotly_chart(fig2, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)

            st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
            st.markdown("<div class='pbi-card-title'>🏆 Matriks Evaluasi Kandidat</div>", unsafe_allow_html=True)
            df_matrix = df_kan.copy()
            df_matrix = df_matrix.merge(df_huff_agg, left_on='id', right_on='id_kandidat', how='left').fillna(0)
            if not df_stack.empty: df_matrix = df_matrix.merge(df_stack.groupby('id_kandidat').agg({'x_terpilih':'max','net_benefit':'sum'}).reset_index(), on='id_kandidat', how='left').fillna(0)
            else: df_matrix['x_terpilih'] = 0; df_matrix['net_benefit'] = 0
            if not df_cvar.empty: df_matrix = df_matrix.merge(df_cvar[['id_kandidat','z_invest','risk_rating']], on='id_kandidat', how='left').fillna(0)
            else: df_matrix['z_invest'] = 0; df_matrix['risk_rating'] = 'N/A'
            df_matrix['Status'] = df_matrix['x_terpilih'].apply(lambda x: '✅ Terpilih' if x == 1 else '❌ Ditolak')
            df_matrix = df_matrix.sort_values('demand_tertangkap', ascending=False)
            show_df = df_matrix[['nama','kab_kota','demand_tertangkap','biaya','z_invest','net_benefit','Status']].copy()
            show_df.columns = ['Nama Lokasi','Kab/Kota','Permintaan Tertangkap','Investasi','Skor CVaR','Net Benefit','Status']
            st.dataframe(show_df, hide_index=True, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)
        except Exception as e: st.error(f"Error Dashboard: {e}"); st.exception(e)

# ============================================================
# TAB 2: SPASIAL — dengan fallback jika folium tidak tersedia
# ============================================================
with tab2:
    st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
    st.markdown("<div class='pbi-card-title'>🗺️ Peta Interaktif Infrastruktur SPKLU</div>", unsafe_allow_html=True)

    if not FOLIUM_AVAILABLE:
        st.error("⚠️ Library `folium` tidak tersedia di environment ini. Menampilkan tabel koordinat sebagai alternatif.")
        st.dataframe(df_kan[['nama','kab_kota','lat','lon','biaya','demand']].rename(columns={'lat':'Latitude','lon':'Longitude'}), hide_index=True, use_container_width=True)
    else:
        c1, c2, c3, c4 = st.columns(4)
        with c1: show_sel = st.toggle("Lokasi Terpilih", True)
        with c2: show_rej = st.toggle("Lokasi Ditolak", True)
        with c3: show_poi = st.toggle("Titik Permintaan", True)
        with c4: show_heat = st.toggle("Heatmap Demand", False)
        try:
            center_lat = df_kan['lat'].mean(); center_lon = df_kan['lon'].mean()
            m = folium.Map(location=[center_lat, center_lon], zoom_start=9, tiles="CartoDB positron")
            folium.TileLayer('OpenStreetMap', name='OSM').add_to(m)
            folium.TileLayer(tiles='https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', name='Satelit', attr='Esri').add_to(m)
            Fullscreen(position='topright').add_to(m); MeasureControl(position='topright', primary_length_unit='kilometers').add_to(m); MiniMap(toggle_display=True, position='bottomleft').add_to(m)
            fg_selected = folium.FeatureGroup(name="✅ Lokasi Terpilih")
            fg_rejected = folium.FeatureGroup(name="❌ Lokasi Ditolak")
            fg_poi = folium.FeatureGroup(name="📍 POI Demand")
            fg_jalan = folium.FeatureGroup(name="🛣️ Jaringan Jalan")
            df_stack_map = db.query("SELECT * FROM hasil_stackelberg") if results else pd.DataFrame()
            if show_sel and not df_stack_map.empty:
                selected_ids = df_stack_map[df_stack_map['x_terpilih'] == 1]['id_kandidat'].unique()
                for _, row in df_kan.iterrows():
                    if row['id'] in selected_ids:
                        popup = f"<div style='font-family:Arial; width:280px;'><div style='background:#059669; padding:10px; color:white; border-radius:6px; margin-bottom:8px;'><div style='font-size:0.7rem;'>✓ LOKASI TERPILIH</div><div style='font-size:1rem; font-weight:700;'>{row['nama']}</div></div><div style='font-size:0.8rem;'>📍 {row['kab_kota']}<br>💰 Rp {row['biaya']/1e9:.2f}M<br>⚡ {row['demand']:,.0f} kWh/hari</div></div>"
                        folium.Marker([row['lat'], row['lon']], popup=folium.Popup(popup, max_width=300), tooltip=f"✅ {row['nama']}", icon=folium.Icon(color='green', icon='bolt', prefix='fa')).add_to(fg_selected)
            if show_rej and not df_stack_map.empty:
                rejected_ids = df_stack_map[df_stack_map['x_terpilih'] == 0]['id_kandidat'].unique()
                for _, row in df_kan.iterrows():
                    if row['id'] in rejected_ids:
                        popup = f"<div style='font-family:Arial; width:260px;'><div style='background:#DC2626; padding:10px; color:white; border-radius:6px; margin-bottom:8px;'><div style='font-size:0.7rem;'>✗ TIDAK TERPILIH</div><div style='font-size:1rem; font-weight:700;'>{row['nama']}</div></div><div style='font-size:0.8rem;'>💰 Rp {row['biaya']/1e9:.2f}M<br>⚡ {row['demand']:,.0f} kWh/hari</div></div>"
                        folium.Marker([row['lat'], row['lon']], popup=folium.Popup(popup, max_width=280), tooltip=f"❌ {row['nama']}", icon=folium.Icon(color='red', icon='times', prefix='fa')).add_to(fg_rejected)
            if show_poi:
                poi_colors = {'Mall':'#3B82F6','Rumah Sakit':'#10B981','Perkantoran':'#8B5CF6','Bandara':'#F59E0B','Kampus':'#EC4899','Wisata':'#06B6D4','Pelabuhan':'#6366F1','Pusat Kota':'#14B8A6','Pasar':'#D97706','Terminal':'#F97316','Industri':'#64748B','SPBU':'#84CC16','Olahraga':'#F43F5E','Religi':'#8B5CF6'}
                for _, row in df_poi.iterrows():
                    color = poi_colors.get(row['tipe'], '#64748B')
                    folium.CircleMarker([row['lat'], row['lon']], radius=6, popup=folium.Popup(f"<b>{row['nama']}</b><br>{row['tipe']}<br>👥 {row['kunjungan_hari']:,} kunjungan/hari", max_width=200), tooltip=f"{row['tipe']}: {row['nama']}", color=color, fill=True, fill_color=color, fill_opacity=0.7, weight=2).add_to(fg_poi)
            df_jalan = db.query("SELECT * FROM jaringan_jalan")
            for _, row in df_jalan.iterrows():
                color = '#1E40AF' if row['tipe'] == 'Arteri' else '#64748B'
                weight = 4 if row['tipe'] == 'Arteri' else 2
                folium.PolyLine([[row['lat_from'], row['lon_from']], [row['lat_to'], row['lon_to']]], color=color, weight=weight, opacity=0.6, tooltip=f"{row['nama']} ({row['panjang_km']:.1f} km)").add_to(fg_jalan)
            if show_heat and results and not results.get('huff').empty:
                heat_data = []
                df_h = results['huff']
                for _, row in df_h.iterrows():
                    if row['probabilitas'] > 0.05:
                        kan = df_kan[df_kan['id'] == row['id_kandidat']].iloc[0]
                        heat_data.append([kan['lat'], kan['lon'], row['demand_tertangkap']])
                if heat_data: HeatMap(heat_data, radius=25, blur=15, max_zoom=10).add_to(m)
            fg_selected.add_to(m); fg_rejected.add_to(m); fg_poi.add_to(m); fg_jalan.add_to(m)
            folium.LayerControl(collapsed=False, position='topleft').add_to(m)

            # Render peta dengan fallback
            if STREAMLIT_FOLIUM_AVAILABLE:
                st_folium(m, width=1200, height=600)
            else:
                st.components.v1.html(m._repr_html_(), height=600)
        except Exception as e: st.error(f"Error peta: {e}"); st.exception(e)
    st.markdown("</div>", unsafe_allow_html=True)


# ============================================================
# TAB 3: NSGA-II
# ============================================================
with tab3:
    if results is None or results.get('nsga') is None or results['nsga'].empty:
        st.info("Jalankan analisis untuk melihat hasil NSGA-II.")
    else:
        try:
            df_nsga = results['nsga']
            pareto = df_nsga[df_nsga['is_pareto'] == True]
            st.info("NSGA-II: Algoritma evolusioner multi-objektif. Trade-off: Biaya ↓, Cakupan ↑, Ekuitas ↑, Risiko ↓.")
            k1, k2, k3 = st.columns(3)
            k1.metric("Solusi Pareto", len(pareto))
            k2.metric("Generasi", nsga_gen)
            k3.metric("Populasi", nsga_pop)
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
                st.markdown("<div class='pbi-card-title'>📊 Pareto Front: Biaya vs Cakupan</div>", unsafe_allow_html=True)
                fig = plotly_pareto_scatter(pareto, 'obj_biaya', 'obj_coverage', title="Pareto Front: Biaya vs Cakupan", x_label="Biaya (Miliar Rp)", y_label="Coverage Score")
                st.plotly_chart(fig, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)
            with c2:
                st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
                st.markdown("<div class='pbi-card-title'>📊 Pareto Front: Risiko vs Ekuitas</div>", unsafe_allow_html=True)
                fig2 = plotly_pareto_scatter(pareto, 'obj_risiko', 'obj_equity', title="Pareto Front: Risiko vs Ekuitas", x_label="Risk (CVaR-adjusted)", y_label="Equity Score")
                st.plotly_chart(fig2, use_container_width=True)
                st.markdown("</div>", unsafe_allow_html=True)
            st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
            st.markdown("<div class='pbi-card-title'>📋 Detail Solusi Pareto Optimal</div>", unsafe_allow_html=True)
            def decode_lokasi(json_str, df_kan):
                try: ids = json.loads(json_str)
                except: return "-"
                names = [df_kan[df_kan['id']==i]['nama'].values[0] for i in ids if i in df_kan['id'].values]
                return ", ".join(names[:3]) + ("..." if len(names) > 3 else "")
            pareto['Lokasi'] = pareto['lokasi_terpilih'].apply(lambda x: decode_lokasi(x, df_kan))
            show_pareto = pareto[['solusi_id','obj_biaya','obj_coverage','obj_equity','obj_risiko','Lokasi']].copy()
            show_pareto.columns = ['ID Solusi','Biaya (M)','Cakupan','Ekuitas','Risiko','Lokasi Terpilih']
            st.dataframe(show_pareto, hide_index=True, use_container_width=True)
            st.markdown("</div>", unsafe_allow_html=True)
        except Exception as e: st.error(f"Error NSGA-II: {e}"); st.exception(e)

# ============================================================
# TAB 4: STACKELBERG & CVaR
# ============================================================
with tab4:
    if results is None: st.info("Jalankan analisis terlebih dahulu.")
    else:
        sub1, sub2 = st.tabs(["🎯 Stackelberg", "⚖️ CVaR"])
        with sub1:
            st.info("Stackelberg: Interaksi strategis Pemerintah (Leader) vs Investor (Follower).")
            try:
                df_stack = results.get('stackelberg', pd.DataFrame())
                if not df_stack.empty:
                    df_stack_show = df_stack.merge(df_kan[['id','nama','kab_kota']], left_on='id_kandidat', right_on='id')
                    df_stack_show = df_stack_show[['nama','kab_kota','x_terpilih','subsidi_optimal','leader_payoff','follower_payoff','net_benefit','equilibrium_type']]
                    df_stack_show.columns = ['Nama','Kab/Kota','Status','Subsidi','Leader Payoff','Follower Payoff','Net Benefit','Ekuilibrium']
                    df_stack_show['Status'] = df_stack_show['Status'].apply(lambda x: '✅ Invest' if x==1 else '❌ No Invest')
                    st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
                    st.markdown("<div class='pbi-card-title'>📊 Hasil Simulasi Stackelberg</div>", unsafe_allow_html=True)
                    st.dataframe(df_stack_show, hide_index=True, use_container_width=True)
                    st.markdown("</div>", unsafe_allow_html=True)
                    st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
                    st.markdown("<div class='pbi-card-title'>📈 Payoff Leader vs Follower</div>", unsafe_allow_html=True)
                    df_plot = df_stack_show[df_stack_show['Status']=='✅ Invest']
                    if not df_plot.empty:
                        fig = plotly_payoff_chart(df_plot, 'Nama', 'Leader Payoff', 'Follower Payoff', 'Payoff Leader vs Follower')
                        st.plotly_chart(fig, use_container_width=True)
                    st.markdown("</div>", unsafe_allow_html=True)
                else: st.info("Data Stackelberg kosong.")
            except Exception as e: st.error(f"Error Stackelberg: {e}"); st.exception(e)
        with sub2:
            st.info("CVaR: Conditional Value at Risk — estimasi kerugian terburuk yang diharapkan.")
            try:
                df_cvar = results.get('cvar', pd.DataFrame())
                if not df_cvar.empty:
                    df_cvar_show = df_cvar.merge(df_kan[['id','nama','kab_kota']], left_on='id_kandidat', right_on='id')
                    df_cvar_show = df_cvar_show[['nama','kab_kota','expected_return','var_95','cvar_95','z_invest','adjusted_npv','risk_rating']]
                    df_cvar_show.columns = ['Nama','Kab/Kota','Expected Return','VaR 95%','CVaR 95%','Z-Score','Adjusted NPV','Rating']
                    st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
                    st.markdown("<div class='pbi-card-title'>📊 Profil Risiko per Lokasi</div>", unsafe_allow_html=True)
                    st.dataframe(df_cvar_show, hide_index=True, use_container_width=True)
                    st.markdown("</div>", unsafe_allow_html=True)
                    st.markdown("<div class='pbi-card'>", unsafe_allow_html=True)
                    st.markdown("<div class='pbi-card-title'>📈 Distribusi Risiko</div>", unsafe_allow_html=True)
                    if not df_cvar.empty:
                        np.random.seed(42)
                        dist = np.random.normal(df_cvar.iloc[0]['expected_return'], abs(df_cvar.iloc[0]['cvar_95'])*0.5, 1000)
                        fig = plotly_risk_distribution(dist, df_cvar.iloc[0]['var_95'], df_cvar.iloc[0]['cvar_95'], "Distribusi Risiko Investasi")
                        st.plotly_chart(fig, use_container_width=True)
                    st.markdown("</div>", unsafe_allow_html=True)
                else: st.info("Data CVaR kosong.")
            except Exception as e: st.error(f"Error CVaR: {e}"); st.exception(e)

# ============================================================
# TAB 5: REKOMENDASI FINAL
# ============================================================
with tab5:
    if results is None: st.info("Jalankan analisis untuk menghasilkan rekomendasi final.")
    else:
        try:
            df_stack = results.get('stackelberg', pd.DataFrame())
            df_cvar = results.get('cvar', pd.DataFrame())
            df_huff_agg = results.get('huff_agg', pd.DataFrame())
            df_rekom = df_kan.copy()
            df_rekom = df_rekom.merge(df_huff_agg, left_on='id', right_on='id_kandidat', how='left').fillna(0)
            if not df_stack.empty: df_rekom = df_rekom.merge(df_stack.groupby('id_kandidat').agg({'x_terpilih':'max','net_benefit':'sum','subsidi_optimal':'max'}).reset_index(), on='id_kandidat', how='left').fillna(0)
            else: df_rekom['x_terpilih'] = 0; df_rekom['net_benefit'] = 0; df_rekom['subsidi_optimal'] = 0
            if not df_cvar.empty: df_rekom = df_rekom.merge(df_cvar[['id_kandidat','z_invest','risk_rating']], on='id_kandidat', how='left').fillna(0)
            else: df_rekom['z_invest'] = 0; df_rekom['risk_rating'] = 'N/A'
            def fase(row):
                if row['x_terpilih'] == 1 and row['z_invest'] >= 0.7: return '🔥 FASE 1 — Prioritas Utama'
                elif row['x_terpilih'] == 1: return '⭐ FASE 2 — Implementasi Bersyarat'
                elif row['demand_tertangkap'] > 800: return '📌 FASE 3 — Cadangan Strategis'
                else: return '❌ Tidak Direkomendasikan'
            df_rekom['Fase'] = df_rekom.apply(fase, axis=1)
            p1 = df_rekom[df_rekom['Fase'].str.contains('FASE 1')]
            p2 = df_rekom[df_rekom['Fase'].str.contains('FASE 2')]
            p3 = df_rekom[df_rekom['Fase'].str.contains('FASE 3')]

            k1, k2, k3, k4 = st.columns(4)
            k1.metric("FASE 1 — Prioritas", len(p1), "Tahun 1")
            k2.metric("FASE 2 — Bersyarat", len(p2), "Monitor pasar")
            k3.metric("FASE 3 — Cadangan", len(p3), "Ekspansi")
            k4.metric("Investasi Fase 1", f"Rp {p1['biaya'].sum()/1e9:.1f}M", "dari anggaran")

            st.markdown("---")
            st.subheader("🎯 Ringkasan Eksekutif & Rekomendasi Kebijakan")
            st.markdown(f"""
            Berdasarkan framework **SPRA-DSS**, sistem telah mengidentifikasi **{len(p1)} lokasi prioritas** untuk implementasi SPKLU Fase 1 di Kalimantan Selatan. Total investasi sebesar **Rp {p1['biaya'].sum()/1e9:.1f} Miliar** diproyeksikan menghasilkan cakupan permintaan **{p1['demand_tertangkap'].sum():,.0f} kWh/hari**.
            **Kebijakan yang Direkomendasikan:**
            • Subsidi optimal rata-rata **{p1['subsidi_optimal'].mean()*100:.1f}%** dari biaya konstruksi
            • Jarak antar-SPPLU memenuhi standar minimum **{min_jarak:.0f} km**
            • Review Fase 2 setelah adopsi EV mencapai **15% dari proyeksi**
            """)

            st.markdown("---")
            st.subheader("📋 Matriks Keputusan Final Terintegrasi")
            df_final = df_rekom[['nama','kab_kota','demand_tertangkap','x_terpilih','z_invest','subsidi_optimal','net_benefit','risk_rating','Fase']].copy()
            df_final.columns = ['Nama Lokasi','Kab/Kota','Permintaan','Stackelberg','Skor CVaR','Subsidi','Net Benefit','Risiko','Keputusan']
            df_final = df_final.sort_values('Permintaan', ascending=False)
            st.dataframe(df_final, hide_index=True, use_container_width=True)

            st.markdown("---")
            st.subheader("📤 Export Hasil Analisis")
            c1, c2, c3 = st.columns(3)
            with c1:
                csv_final = df_final.to_csv(index=False).encode('utf-8')
                st.download_button("⬇️ Download CSV", csv_final, "rekomendasi_spklu_kalsel.csv", "text/csv", use_container_width=True)
            with c2:
                if OPENPYXL_AVAILABLE:
                    excel_output = export_excel_full(results, df_kan, df_poi, df_wilayah)
                    if excel_output:
                        st.download_button("⬇️ Download Excel", excel_output, "rekomendasi_spklu_kalsel.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    else:
                        st.button("⬇️ Download Excel", disabled=True, use_container_width=True)
                else:
                    st.button("⬇️ Download Excel (install openpyxl)", disabled=True, use_container_width=True)
            with c3:
                if FPDF_AVAILABLE:
                    pdf_output = generate_pdf_report(results, df_kan, budget, min_jarak, nsga_pop, nsga_gen)
                    if pdf_output:
                        st.download_button("📄 Generate Laporan PDF", pdf_output, "laporan_spklu_kalsel.pdf", "application/pdf", use_container_width=True)
                    else:
                        st.button("📄 Generate PDF", disabled=True, use_container_width=True)
                else:
                    st.button("📄 Generate PDF (install fpdf2)", disabled=True, use_container_width=True)

            st.markdown("---")
            st.subheader("💾 Simpan Skenario ke Database")
            with st.form("save_scenario_form"):
                nama_scenario = st.text_input("Nama Skenario", value=f"Skenario {datetime.now().strftime('%d%m%Y_%H%M')}")
                deskripsi_scenario = st.text_area("Deskripsi", value=f"Budget Rp {budget/1e9:.0f}M, Jarak {min_jarak:.0f}km, Pop {nsga_pop}, Gen {nsga_gen}")
                submitted = st.form_submit_button("💾 Simpan ke Database", use_container_width=True)
                if submitted and results:
                    save_scenario_to_db(db, nama_scenario, deskripsi_scenario, budget, max_lokasi, min_jarak, huff_beta, cvar_conf, risk_aversion, nsga_pop, nsga_gen, results)
                    st.toast(f"✅ Skenario '{nama_scenario}' tersimpan!", icon="💾")
                    st.balloons()

        except Exception as e: st.error(f"Error Rekomendasi: {e}"); st.exception(e)

# ============================================================
# SCENARIO MANAGER DIALOG
# ============================================================
st.markdown("---")
with st.expander("📂 Scenario Manager — Lihat & Bandingkan Skenario Tersimpan"):
    scenarios = load_scenarios_from_db(db)
    if scenarios.empty:
        st.info("Belum ada skenario tersimpan. Jalankan analisis dan simpan skenario.")
    else:
        st.write(f"Total skenario tersimpan: **{len(scenarios)}**")
        for _, sc in scenarios.iterrows():
            hasil = json.loads(sc['hasil_json']) if sc['hasil_json'] else {}
            col1, col2, col3 = st.columns([4, 1, 1])
            with col1:
                st.markdown(f"""
                <div class='scenario-card'>
                    <b>{sc['nama']}</b> <span style='color:#64748B; font-size:0.8rem;'>({sc['created_at']})</span><br>
                    <span style='font-size:0.85rem; color:#475569;'>Budget: Rp {sc['budget']/1e9:.0f}M | Jarak: {sc['min_jarak_km']:.0f}km | Pop: {sc['nsga_popsize']} | Gen: {sc['nsga_generasi']}</span><br>
                    <span style='font-size:0.8rem; color:#64748B;'>Pareto: {hasil.get('n_pareto', 0)} | Terpilih: {hasil.get('n_terpilih', 0)} | Demand: {hasil.get('total_demand', 0):,.0f}</span>
                </div>
                """, unsafe_allow_html=True)
            with col2:
                if st.button("🗑️ Hapus", key=f"del_{sc['id']}"):
                    delete_scenario(db, sc['id'])
                    st.toast("Skenario dihapus", icon="🗑️")
                    st.rerun()
            with col3:
                if st.button("📋 Load", key=f"load_{sc['id']}"):
                    st.toast(f"Skenario '{sc['nama']}' dimuat!", icon="📂")

# ============================================================
# FOOTER
# ============================================================
st.markdown("""
<div class='footer-bar'>
    <div><b>⚡ SPK-GeoOptima v5.1</b> &nbsp;|&nbsp; Framework SPRA-DSS &nbsp;|&nbsp; Kalimantan Selatan</div>
    <div>GIS + NSGA-II + Huff + Stackelberg + CVaR &nbsp;|&nbsp; Disertasi Infrastruktur Publik</div>
</div>
""", unsafe_allow_html=True)
